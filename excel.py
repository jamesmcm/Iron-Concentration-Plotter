printing = 0
from openpyxl.reader.excel import load_workbook

def getValues(filename):


    try:
        wb = load_workbook(filename)
        ws=wb.get_sheet_by_name("User")
        resistivity=(ws.cell('C6')).value
        ws = wb.get_sheet_by_name("Calc")
        tauvalues=[]
        deltanvalues=[]
        #tau values full131
        #Want to cut off effects from trapping, so cut off at minimum tau value
        for row in ws.range('L15:L131'):
            for cell in row:
                tauvalues.append(float(cell.value))
        if printing == 1:
            print "-----"
        #delta n values
        for row in ws.range('S15:S131'):
            for cell in row:
                deltanvalues.append(float(cell.value))
        #cannot use global minimum as want to find first local minimum
        limit=getlocalmin(tauvalues)

        returnlist=[resistivity, tauvalues, deltanvalues]
        return returnlist
    except:
        print "Failed to open workbook"
        return [0,0,0]

   

def getlocalmin(tauvalues):
    i=0
    while i<(len(tauvalues)-4):
        if tauvalues[i+5]>tauvalues[i]:
            #3 seems like a good compromise here
            #perhaps make this editable in GUI
            break
        i+=1
    return i

if __name__ == "__main__":
    import sys
    printing=1
    listlist=getValues(str(sys.argv[1]))
    resistivity=listlist[0]
    tauvalues=listlist[1]
    deltanvalues=listlist[2]
    i=0
    print "Resisitivity: " + str(resistivity)
    print "Tau    DeltaN"
    while i<len(tauvalues):
        print str(tauvalues[i]) + "    " + str(deltanvalues[i]) 
        i+=1


