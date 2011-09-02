import matplotlib 
matplotlib.use('GTK')  
import pygtk
pygtk.require("2.0")
import gtk
import excel
from numpy import savetxt, loadtxt
import numpy as np
from scipy.interpolate import interp1d
import os
from matplotlib.pyplot import plot, savefig, rc
from matplotlib.colorbar import Colorbar, ColorbarBase
from matplotlib.figure import Figure   
from matplotlib.ticker import LogLocator, LogFormatter 
import matplotlib.cm as cm
from math import pow
from matplotlib.axes import Subplot   
from matplotlib.backends.backend_gtkagg import NavigationToolbar2GTKAgg as NavigationToolbar
from matplotlib.backends.backend_gtkagg import FigureCanvasGTKAgg  

import doping
import bildregistrierung_ng as bildreg
from openpyxl.reader.excel import load_workbook
import concentration
from matplotlib.colors import LogNorm
from statlib import stats

import warnings
warnings.simplefilter('ignore')

#matplotlib.rc('text', usetex=True) doesn't work

class MyApp(object):       
	def __init__(self):
	    self.builder = gtk.Builder()
	    self.builder.add_from_file("myapp.xml")
	    self.builder.connect_signals({ "on_MainWindow_destroy" : gtk.main_quit, "on_menuquit_activate" : gtk.main_quit, "on_filebeforebtn_file_set" : self.plotgraph1, "on_ErrorWindow_close" : self.errorclose, "scaletxtedit":self.scaletxtedit, "scalebtnclicked":self.scalebtnclicked, "resetbtnclicked":self.plotgraph1, "resistivityresponse" : self.errorclose, "ironcalcbtnclicked" : self.ironcalcbtnclicked, "scalebtn2clicked":self.scalebtn2clicked, "resetbtn2clicked":self.resetbtn2clicked, "saveplotbtnclicked":self.saveplotbtnclicked, "plotdialogcancelbtnclicked":self.plotdialogcancelbtnclicked, "plotdialogsavebtnclicked":self.plotdialogsavebtnclicked, "savedatabtnclicked":self.savedatabtnclicked, "datadialogsavebtnclicked":self.datadialogsavebtnclicked, "datadialogcancelbtnclicked":self.datadialogcancelbtnclicked, "cfactorbtnclicked":self.cfactorbtnclicked, "plotflippeddialogclose":self.errorclose, "restoredefaultsbtnclicked":self.restoredefaultsbtnclicked, "plcalcbtnclicked":self.plcalcbtnclicked, "plfileset":self.plfileset, "plplotallbtnclicked":self.plplotallbtnclicked, "zoominbtnclicked":self.zoominbtnclicked, "zoomoutbtnclicked":self.zoomoutbtnclicked, "panleftbtnclicked":self.panleftbtnclicked, "panrightbtnclicked":self.panrightbtnclicked, "panupbtnclicked":self.panupbtnclicked, "pandownbtnclicked":self.pandownbtnclicked, "plplotdialogsavebtnclicked": self.plplotdialogsavebtnclicked, "plplotdialogcancelbtnclicked":self.plplotdialogcancelbtnclicked, "savefemapbtnclicked":self.savefemapbtnclicked, "overwriteresponseclicked":self.overwriteresponseclicked, "numbify":self.numbify, "pl1genlevelokbtnclicked":self.pl1genlevelokbtnclicked, "pl1genlevelcancelbtnclicked":self.pl1genlevelcancelbtnclicked, "getfebtnclicked":self.getfebtnclicked, "whichsavefebtnclicked":self.whichsavefebtnclicked, "whichsavelifebtnclicked":self.whichsavelifebtnclicked, "whichsavecancelbtnclicked":self.whichsavecancelbtnclicked, "editcolorbarbtnclicked":self.editcolorbarbtnclicked, "editcolorbarokbtnclicked":self.editcolorbarokbtnclicked, "editcolorbarcancelbtnclicked":self.editcolorbarcancelbtnclicked, "saveirondatabtnclicked":self.saveirondatabtnclicked, "recalccopbtnclicked":self.recalccopbtnclicked, "restoredopingbtnclicked":self.restoredopingbtnclicked, "attemptcorrectionbtnclicked":self.attemptcorrectionbtnclicked, "donothingbtnclicked":self.donothingbtnclicked, "saveadjusteddatabtnclicked":self.saveadjusteddatabtnclicked, "calcdopefromresbtnclicked":self.calcdopefromresistivity, "resbeforevalbtnclicked":self.resbeforevalbtnclicked, "resaftervalbtnclicked":self.resaftervalbtnclicked, "resmyvalbtnclicked":self.resmyvalbtnclicked})
	    self.window = self.builder.get_object("MainWindow")
	    #self.window.fullscreen()
	    filter1 = gtk.FileFilter()
	    filter1.set_name("XLSM")
	    filter1.add_pattern("*.xlsm")
	    filter2 = gtk.FileFilter()
	    filter2.set_name("TXT")
	    filter2.add_pattern("*.txt")
	    filterplot = gtk.FileFilter()
	    filterplot.set_name("PNG")
	    filterplot.add_pattern("*.png")
	    self.builder.get_object("plotfilesavedialog").add_filter(filterplot)
	    self.builder.get_object("plplotfilesavedialog").add_filter(filterplot)
	    self.builder.get_object("plbeforefile").add_filter(filter2)
	    self.builder.get_object("plafterfile").add_filter(filter2)
	    self.builder.get_object("datafilesavedialog").add_filter(filter2)
	    
	    self.beforefileselect=self.builder.get_object("filebeforebtn")
	    self.beforefileselect.add_filter(filter1)
	    self.afterfileselect=self.builder.get_object("fileafterbtn")
	    self.afterfileselect.add_filter(filter1)
	    self.graph1counter=0
	    self.ironcounter=0
	    self.adjusted=0
	    self.plcounter=0
	    self.cursave=""
	    self.curid=None

	    self.combobox = gtk.combo_box_new_text()
	    self.builder.get_object("combospace").add(self.combobox)
	    self.builder.get_object("combospace").show_all()
	    self.combobox.append_text('Gray')
	    self.combobox.append_text('Jet')
	    self.combobox.append_text('Hot')
	    self.combobox.set_active(0)


	    self.linbutton = gtk.RadioButton(None, "Linear")
	    self.linbutton.connect("toggled", self.buttontoggle, "linear")
	    self.builder.get_object("rbutton1vbox").pack_start(self.linbutton, True, True, 0)
	    self.linbutton.show()
	    self.logbutton = gtk.RadioButton(self.linbutton, "Base-10 Logarithmic")
	    self.logbutton.connect("toggled", self.buttontoggle, "logarithmic")
	    self.builder.get_object("rbutton2vbox").pack_start(self.logbutton, True, True, 0)
	    self.logbutton.show()
	    self.linbutton.set_active(True)
	    #self.logbutton.set_active(False)
	    self.lifemaptype="linear"
	    self.templifemaptype="linear"

	    #iron radio buttons
	    self.felinbutton = gtk.RadioButton(None, "Linear")
	    self.felinbutton.connect("toggled", self.febuttontoggle, "linear")
	    self.builder.get_object("ironradio1vbox").pack_start(self.felinbutton, True, True, 0)
	    self.felinbutton.show()
	    self.felogbutton = gtk.RadioButton(self.felinbutton, "Base-10 Logarithmic")
	    self.felogbutton.connect("toggled", self.febuttontoggle, "logarithmic")
	    self.builder.get_object("ironradio2vbox").pack_start(self.felogbutton, True, True, 0)
	    self.felogbutton.show()
	    self.felinbutton.set_active(True)
	    self.ironmaptype="linear"
	    self.tempironmaptype="linear"
	    self.plconid=self.builder.get_object("statusbar").get_context_id("graphs")
	    self.constants=[1.1E7,1.3E-14,1.28E13,7E-17,5E-15,1.22571E15,3E-15]
	    self.window.show()
	    cm.gray.set_under('r') #add these to colorbar somehow?
	    #cm.gray.set_over('b')
	    #cm.gray.set_bad('g')
	    cm.jet.set_under('w') 
	   #cm.jet.set_over('k')
	    #cm.hot.set_under('b')
	    #cm.hot.set_over('g')
	    self.cmap=cm.gray
	    
	def plotgraph1(self, widget):
		befile=self.builder.get_object("filebeforebtn")
		affile=self.builder.get_object("fileafterbtn")
		graphview = self.builder.get_object("graph1")  
		#remember the order of the constants
		#vthermal sigmani p1i sigmapi sigmanb n1b sigmapb
		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]
		if self.graph1counter==0:
			self.figure1 = Figure(figsize=(6,4), dpi=72)  
			self.axis1 = self.figure1.add_subplot(111)
		if self.graph1counter>0:
			self.axis1 = self.figure1.add_subplot(111) 
			self.axis1.clear()
			graphview.remove(self.canvas1)
			self.builder.get_object("toolbar1").remove(self.toolbar1)	
			
		if befile.get_filename()!=None:
			self.plot1(befile, "ro", "b-", "Before Illumination")
		if affile.get_filename()!=None:
			self.plot1(affile, "go", "r-", "After Illumination")
			
		self.axis1.set_xlabel('Minority Carrier Density (cm$^{-3}$)')   
		self.axis1.set_ylabel('Measured Lifetime (s)')   
		self.axis1.ticklabel_format(style='sci',scilimits=(0,0))
		self.axis1.grid(True)
		self.canvas1=FigureCanvasGTKAgg(self.figure1)
		self.canvas1.show()
		graphview.pack_start(self.canvas1, True, True)
		self.toolbar1 = NavigationToolbar(self.canvas1, self.builder.get_object("MainWindow"))
		self.builder.get_object("toolbar1").pack_start(self.toolbar1, False, False)
		self.scaletxtedit(widget)
		self.graph1counter+=1
		if befile.get_filename()!=None and affile.get_filename()!=None:
			self.checkresistivity(widget)
			self.builder.get_object("plbeforefile").set_sensitive(True)
			self.builder.get_object("recalccopbtn").set_sensitive(True)
			self.builder.get_object("calcdopefromres").set_sensitive(True)
			self.builder.get_object("restoredopingbtn").set_sensitive(True)

	def plot1(self, file1, pointcol, linecol, mylabel):
		#note there is no adjustment for extreme values here, although an adjustment of the limit method could perhaps provide this
		filename=file1.get_filename()
		fitpoints=[]
		datalist=excel.getValues(filename)
		if datalist==[0,0,0]:
			self.errorshow(self)
			return 1
		limit = excel.getlocalmin(datalist[1])
		self.axis1.plot(datalist[2],datalist[1], pointcol)
		self.axis1.plot((datalist[2])[0:limit+1],(datalist[1])[0:limit+1], linecol, label=mylabel)
		self.axis1.plot(datalist[2],datalist[1], pointcol)		
		self.axis1.legend()
	def checkresistivity(self,widget):
		#remember the order of the constants
		#vthermal sigmani p1i sigmapi sigmanb n1b sigmapb
		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]
		befile=self.builder.get_object("filebeforebtn")
		affile=self.builder.get_object("fileafterbtn")
		belist=excel.getValues(befile.get_filename())
		aflist=excel.getValues(affile.get_filename())
		self.oldres=belist[0]
		self.dope=doping.calcDoping(self.oldres)
		self.olddope=self.dope
		self.builder.get_object("crossovertheory").set_text("%.4g" % concentration.COPcalc(self.constants, self.dope))
		self.builder.get_object("doping").set_text("%.4g" % self.dope)
		self.builder.get_object("pldopingtxt").set_text("%.4g" % self.dope)		
		if belist[0]!=aflist[0]:
			self.resbeforeval=belist[0]
			self.resafterval=aflist[0]
			self.builder.get_object("resbeforevallabel").set_label("%.4g" % self.resbeforeval)
			self.builder.get_object("resaftervallabel").set_label("%.4g" % self.resafterval)
			self.builder.get_object("reswindow").show()

		self.builder.get_object("resistivity").set_text("%.4g" % belist[0])

	def resbeforevalbtnclicked(self, widget):
		self.oldres=self.resbeforeval
		self.dope=doping.calcDoping(self.resbeforeval)
		self.builder.get_object("resistivity").set_text("%.4g" % self.resbeforeval)
		self.builder.get_object("pldopingtxt").set_text("%.4g" % self.dope)
		self.builder.get_object("doping").set_text("%.4g" % self.dope)
		self.builder.get_object("crossovertheory").set_text("%.4g" % concentration.COPcalc(self.constants, self.dope))		
		self.builder.get_object("reswindow").hide()

	def resaftervalbtnclicked(self, widget):
		self.dope=doping.calcDoping(self.resafterval)
		self.oldres=self.resafterval
		self.builder.get_object("resistivity").set_text("%.4g" % self.resafterval)
		self.builder.get_object("pldopingtxt").set_text("%.4g" % self.dope)
		self.builder.get_object("doping").set_text("%.4g" % self.dope)
		self.builder.get_object("crossovertheory").set_text("%.4g" % concentration.COPcalc(self.constants, self.dope))		
		self.builder.get_object("reswindow").hide()

	def resmyvalbtnclicked(self, widget):
		myres=float(self.builder.get_object("myrestxt").get_text())
		self.oldres=myres
		self.dope=doping.calcDoping(myres)
		self.builder.get_object("resistivity").set_text("%.4g" % myres)
		self.builder.get_object("pldopingtxt").set_text("%.4g" % self.dope)
		self.builder.get_object("doping").set_text("%.4g" % self.dope)	
		self.builder.get_object("crossovertheory").set_text("%.4g" % concentration.COPcalc(self.constants, self.dope))	
		self.builder.get_object("reswindow").hide()


	def recalccopbtnclicked(self, widget):
		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]
		self.dope=float(self.builder.get_object("doping").get_text())
		self.builder.get_object("crossovertheory").set_text("%.4g" % concentration.COPcalc(self.constants, self.dope))
		self.builder.get_object("pldopingtxt").set_text("%.4g" % self.dope)

	def restoredopingbtnclicked(self, widget):
		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]
		self.dope=self.olddope
		self.builder.get_object("crossovertheory").set_text("%.4g" % concentration.COPcalc(self.constants, self.dope))
		self.builder.get_object("pldopingtxt").set_text("%.4g" % self.dope)
		self.builder.get_object("doping").set_text("%.4g" % self.dope)
		self.builder.get_object("resistivity").set_text("%.4g" % self.oldres)

	def errorclose(self, widget, response):
		widget.hide()
		#response is -7 for close
	
	def errorshow(self, widget):
		errorwin=self.builder.get_object("ErrorWindow")
		errorwin.show()

	def numbify(self, entry):
		text = entry.get_text().strip()

		entry.set_text(''.join([i for i in text if i in '.-eE0123456789']))

	        if gtk.Buildable.get_name(entry)=="doping":
		 	self.builder.get_object("pldopingtxt").set_text("%.4g" % float(entry.get_text()))

	def calcdopefromresistivity(self, widget):
		self.dope=doping.calcDoping(float(self.builder.get_object("resistivity").get_text()))
		self.builder.get_object("pldopingtxt").set_text("%.4g" % self.dope)
		self.builder.get_object("doping").set_text("%.4g" % self.dope)
		
	def scaletxtedit(self, widget):

		if widget.get_name()=="GtkEntry":
			self.numbify(widget)

		if (self.builder.get_object("filebeforebtn").get_filename()!=None or self.builder.get_object("fileafterbtn").get_filename()!=None):
			self.builder.get_object("scalebtn").set_sensitive(True)
		else:
			self.builder.get_object("scalebtn").set_sensitive(False)
		
		if self.builder.get_object("filebeforebtn").get_filename()!=None or self.builder.get_object("fileafterbtn").get_filename()!=None:
			self.builder.get_object("resetbtn").set_sensitive(True)
		else:
			self.builder.get_object("resetbtn").set_sensitive(False)

		if self.builder.get_object("filebeforebtn").get_filename()!=None and self.builder.get_object("fileafterbtn").get_filename()!=None:
			self.builder.get_object("ironcalcbtn").set_sensitive(True)
			self.builder.get_object("cfactorbtn").set_sensitive(True)
		else:
			self.builder.get_object("ironcalcbtn").set_sensitive(False)
			self.builder.get_object("cfactorbtn").set_sensitive(False)

		if self.builder.get_object("plbeforefile").get_filename()!=None and self.builder.get_object("plafterfile").get_filename()!=None  and self.builder.get_object("filebeforebtn").get_filename()!=None and self.builder.get_object("fileafterbtn").get_filename()!=None:
			self.builder.get_object("plcalcbtn").set_sensitive(True)
		else:
			self.builder.get_object("plcalcbtn").set_sensitive(False)

	def scalebtnclicked(self, widget):
		graphview=self.builder.get_object("graph1")

		#This sometimes causes some strange error asking for 2-tuples, it might be a bug in Cygwin

		xmin1=self.builder.get_object("xmin1").get_text()
		xmax1=self.builder.get_object("xmax1").get_text()
		ymin1=self.builder.get_object("ymin1").get_text()
		ymax1=self.builder.get_object("ymax1").get_text()
		xaxis=self.axis1.get_xlim()
		yaxis=self.axis1.get_ylim()
		xmaxmag="%e" % xaxis[1]		
		ymaxmag="%e" % yaxis[1]		
		xmaxmag=xmaxmag[xmaxmag.index("e")+1:len(xmaxmag)]
		ymaxmag=ymaxmag[ymaxmag.index("e")+1:len(ymaxmag)]

		if xmin1 != "" and xmin1 !=None:
			if xmin1.find("e")==-1 and xmin1.find("E")==-1:
				xmin1=float(xmin1)*pow(10, float(xmaxmag))
			self.axis1.set_xlim(left=float(xmin1))

		if xmax1 != "" and xmax1 != None:
			if xmax1.find("e")==-1 and xmax1.find("E")==-1:
				xmax1=float(xmax1)*pow(10, float(xmaxmag))
			self.axis1.set_xlim(right=float(xmax1))

		if ymax1 != "" and ymax1 != None:
			if ymax1.find("e")==-1 and ymax1.find("E")==-1:
				ymax1=float(ymax1)*pow(10, float(ymaxmag))
			self.axis1.set_ylim(top=float(ymax1))
			
		if ymin1 != "" and ymin1 != None:
			if ymin1.find("e")==-1 and ymin1.find("E")==-1:
				ymin1=float(ymin1)*pow(10, float(ymaxmag))
			self.axis1.set_ylim(bottom=float(ymin1))

		yaxis=self.axis1.get_ylim()
		xaxis=self.axis1.get_xlim()
		if yaxis[0]>yaxis[1] or xaxis[0]>xaxis[1]:
			self.builder.get_object("plotflippeddialog").show()

		graphview.remove(self.canvas1)	
		graphview.pack_start(self.canvas1, True, True)

	# def ironcheckbuttons(self, widget):


	# 	if widget.get_name()=="GtkEntry":
	# 		self.numbify(widget)

#I'm not sure if this function is still necessary will check what uses it later.

	def scalebtn2clicked(self, widget):
		graphview=self.builder.get_object("irongraph")
		xmin2=self.builder.get_object("xmin2").get_text()
		xmax2=self.builder.get_object("xmax2").get_text()
		ymin2=self.builder.get_object("ymin2").get_text()
		ymax2=self.builder.get_object("ymax2").get_text()
		xaxis=self.axis2.get_xlim()
		yaxis=self.axis2.get_ylim()
		xmaxmag="%e" % xaxis[1]		
		ymaxmag="%e" % yaxis[1]		
		xmaxmag=xmaxmag[xmaxmag.index("e")+1:len(xmaxmag)]
		ymaxmag=ymaxmag[ymaxmag.index("e")+1:len(ymaxmag)]
		
		if xmax2!="" and xmax2 != None:
			if xmax2.find("e")==-1 and xmax2.find("E")==-1:
				xmax2=float(xmax2)*pow(10, float(xmaxmag))
			self.axis2.set_xlim(right=float(xmax2))
		if xmin2!="" and xmin2 != None:
			if xmin2.find("e")==-1 and xmin2.find("E")==-1:
				xmin2=float(xmin2)*pow(10, float(xmaxmag))
			self.axis2.set_xlim(left=float(xmin2))
		if ymax2!="" and ymax2!=None:
			if ymax2.find("e")==-1 and ymax2.find("E")==-1:
				ymax2=float(ymax2)*pow(10, float(ymaxmag))
			self.axis2.set_ylim(top=float(ymax2))
		if ymin2!="" and ymin2!=None:
			if ymin2.find("e")==-1 and ymin2.find("E")==-1:
				ymin2=float(ymin2)*pow(10, float(ymaxmag))
			self.axis2.set_ylim(bottom=float(ymin2))

		xaxis=self.axis2.get_xlim()
		yaxis=self.axis2.get_ylim()

		if xaxis[0]>xaxis[1] or yaxis[0]>yaxis[1]:
			self.builder.get_object("plotflippeddialog").show()





		graphview.remove(self.canvas2)	
		graphview.pack_start(self.canvas2, True, True)				

	def ironcalcbtnclicked(self, widget):
		self.dope=float(self.builder.get_object("doping").get_text())
		beforelist=excel.getValues(self.builder.get_object("filebeforebtn").get_filename())
		afterlist=excel.getValues(self.builder.get_object("fileafterbtn").get_filename())
		if len(afterlist[1])!=len(afterlist[2]) or len(beforelist[1])!=len(beforelist[2]) or len(afterlist[1])!=len(beforelist[1]):
			print "ERROR"
			#add proper error handling later
		beforelist[2].reverse()
		afterlist[2].reverse()
		beforelist[1].reverse()
		afterlist[1].reverse()
		dictlist=concentration.interpolation(beforelist, afterlist)
		#dictlist[0] is before, dictlist[1] is after values

		for key in dictlist[0].keys():
			if dictlist[1].has_key(key)!=True:
				del dictlist[0][key]

		for key in dictlist[1].keys():
			if dictlist[0].has_key(key)!=True:
				del dictlist[1][key]

		if len(dictlist[0])!=len(dictlist[1]):
			print "dict lengths not equal"
			#add proper error handling later
			#difference is expected but must be handled
		
		ironvalues=[]


		sortedbkeylist=sorted(dictlist[0].iterkeys())
		sortedakeylist=sorted(dictlist[1].iterkeys())
		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]
		#fitvalues=[]
		#remember the order of the constants
		#vthermal sigmani p1i sigmapi sigmanb n1b sigmapb
		#p1b=94310.31142

		for bkey in sortedbkeylist:
			iron=1E-6*(concentration.calcFeConc(self.constants, self.dope, bkey, dictlist[0][bkey], dictlist[1][bkey]))
			ironvalues.append(iron)
			#tn0=1/((99.99/100)*iron*self.constants[4]*self.constants[0])
			#tp0=1/((99.99/100)*iron*self.constants[6]*self.constants[0])
			print "iron: %.4g, tn0: %.4g, tp0: %.4g, doping: %.4g, tbefore: %.4g, tafter: %.4g, deltan: %.4g \n" % (iron, tn0, tp0, self.dope, dictlist[0][bkey], dictlist[1][bkey], bkey)

			#fitvalues.append(((tp0*(self.constants[5]+bkey))+(tn0*(self.dope+p1b+bkey)))/(self.dope+bkey))
		xlim=self.axis1.get_xlim()
		ylim=self.axis1.get_ylim()
		#self.axis1.plot(sortedbkeylist,fitvalues, "m-")
		self.axis1.set_xlim(xlim)
		self.axis1.set_ylim(ylim)
		#plot ironvalues
		graphview = self.builder.get_object("irongraph")  
		if self.ironcounter==0:
			self.figure2 = Figure(figsize=(6,4), dpi=72)  
			self.axis2 = self.figure2.add_subplot(111)
		if self.ironcounter>0:
			self.axis2 = self.figure2.add_subplot(111) 
			self.axis2.clear()
			graphview.remove(self.canvas2)
			self.builder.get_object("toolbar2").remove(self.toolbar2)
		self.axis2.plot(sortedbkeylist,ironvalues, "ro")
		#self.axis2.plot(sortedbkeylist, ironvalues, "b-") #maybe make this within limit later		 
		self.irondata=[sortedbkeylist, ironvalues]
		self.axis2.set_xlabel('Minority Carrier Density (cm$^{-3}$)')   
		self.axis2.set_ylabel('Interstitial Iron Concentration (cm$^{-3}$)') #check units are correct   
		self.axis2.ticklabel_format(style='sci',scilimits=(0,0))
		self.axis2.grid(True)
		self.canvas2=FigureCanvasGTKAgg(self.figure2)
		self.canvas2.show()
		graphview.pack_start(self.canvas2, True, True)
		self.toolbar2 = NavigationToolbar(self.canvas2, self.builder.get_object("MainWindow"))
		self.builder.get_object("toolbar2").pack_start(self.toolbar2, False, False)
		self.ironcounter+=1
		self.builder.get_object("resetbtn2").set_sensitive(True)

		cop=(sortedbkeylist[ironvalues.index(max(ironvalues))] + sortedbkeylist[ironvalues.index(min(ironvalues))])/2
		self.builder.get_object("crossovertxt").set_text("%.4g" % cop)
		self.builder.get_object("saveplotbtn").set_sensitive(True)
		self.builder.get_object("savedatabtn").set_sensitive(True)
		self.builder.get_object("getfebtn").set_sensitive(True)
		self.builder.get_object("meanlabel").set_label("Mean of the interstitial\nIron Concentration from\nthe last 10 values (cm<sup>-3</sup>): ")	
		ironmean=stats.mean(ironvalues[-10:-1])
		self.builder.get_object("meanconctxt").set_text("%.4g" % ironmean)

		#don't delete just change default view
		ironmin=max(ironvalues)
		i=ironvalues.index(ironmin)
		stddevfe=stats.stdev(ironvalues)
		while i<len(ironvalues):
			#0.3 seems good compromise factor
			if abs(ironvalues[i])< abs(ironmean+(0.3*stats.stdev(ironvalues))):
				ironmin=ironvalues[i]
				break
			i+=1
			
		#set default view limits
		self.axis2.set_ylim(bottom=1E10)
		self.axis2.set_ylim(top=ironmin)

		self.builder.get_object("feilabel").set_label("[Fe<sub>i</sub>] (cm<sup>-3</sup>): ")
		self.builder.get_object("getfebtnlabel").set_label("Get [Fe<sub>i</sub>]")
		self.curplot="Fe"
		self.builder.get_object("deltangettxt").set_text("")
		self.builder.get_object("fegettxt").set_text("")
		self.builder.get_object("scalebtn2").set_sensitive(True)	

	def saveplotbtnclicked(self, widget):
		self.builder.get_object("plotfilesavedialog").show()

	def plotdialogcancelbtnclicked(self,widget):
		self.builder.get_object("plotfilesavedialog").hide()
	def plotdialogsavebtnclicked(self, widget):
		self.cursave="plotdialogsave"
		filename=self.builder.get_object("plotfilesavedialog").get_filename()
		if filename[-4:len(filename)]!=".png" and filename[-4:len(filename)]!=".PNG":
			filename += ".png"
		self.currentfilename=filename

		if os.path.exists(filename)==True:
			self.builder.get_object("overwritewarning").show()
		else:
			self.figure2.savefig(filename, format="png")
		#add error handling
		#check if filename has extension, if not append PNG - maybe ask?

		# note that this inherently uses current scaling options - WYSIWYG
		
	def savedatabtnclicked(self,widget):
		self.builder.get_object("datafilesavedialog").show()
		self.datasave="plot"

	def datadialogcancelbtnclicked(self,widget):
		self.builder.get_object("datafilesavedialog").hide()	
	
	def datadialogsavebtnclicked(self,widget):
		filename=self.builder.get_object("datafilesavedialog").get_filename()

		if filename[-4:len(filename)]!=".txt" and filename[-4:len(filename)]!=".TXT":
			filename += ".txt"
		self.currentfilename=filename 
		if self.datasave=="plot":
			self.cursave="feplotdata"
		if self.datasave=="adjusteddata":
			self.cursave="adjusteddata"
		elif self.datasave=="map":
			self.cursave="femapdata"
		if os.path.exists(filename)==True:
			self.builder.get_object("overwritewarning").show()
		else:
			if self.datasave=="plot":
				self.datatxtsave()
			elif self.datasave=="map":
				savetxt(self.currentfilename, self.ironconcmatrix, fmt="%12.6G")
			elif self.datasave=="adjusteddata":
				savetxt(self.currentfilename, self.tauafter, fmt="%12.6G")

		self.builder.get_object("datafilesavedialog").hide()	
	
	
		#add error handling

	def datatxtsave(self):
		self.builder.get_object("datafilesavedialog").hide()
		textfile=open(self.currentfilename, "w")
		textfile.write("deltaN(cm^-3)\t[Fe](cm^-3)\n")
		i=0
		while i<len(self.irondata[0]):
			textfile.write("%e\t%e\n" % (self.irondata[0][i], self.irondata[1][i]))
			i+=1
		textfile.close()

		
	def cfactorbtnclicked(self, widget):
		self.dope=float(self.builder.get_object("doping").get_text())
		beforelist=excel.getValues(self.builder.get_object("filebeforebtn").get_filename())
		afterlist=excel.getValues(self.builder.get_object("fileafterbtn").get_filename())
		if len(afterlist[1])!=len(afterlist[2]) or len(beforelist[1])!=len(beforelist[2]) or len(afterlist[1])!=len(beforelist[1]):
			print "ERROR"
			#add proper error handling later
		beforelist[2].reverse()
		afterlist[2].reverse()
		beforelist[1].reverse()
		afterlist[1].reverse()
		dictlist=concentration.interpolation(beforelist, afterlist)
		#dictlist[0] is before, dictlist[1] is after values

		for key in dictlist[0].keys():
			if dictlist[1].has_key(key)!=True:
				del dictlist[0][key]

		for key in dictlist[1].keys():
			if dictlist[0].has_key(key)!=True:
				del dictlist[1][key]

		if len(dictlist[0])!=len(dictlist[1]):
			print "dict lengths not equal"
			#add proper error handling later
			#difference is expected but must be handled
		
		cvalues=[]

		sortedbkeylist=sorted(dictlist[0].iterkeys())
		sortedakeylist=sorted(dictlist[1].iterkeys())
		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]
		for bkey in sortedbkeylist:
			cvalues.append(concentration.calcPrefactor(self.constants, self.dope, bkey))

		#plot cvalues
		graphview = self.builder.get_object("irongraph")  
		if self.ironcounter==0:
			self.figure2 = Figure(figsize=(6,4), dpi=72)  
			self.axis2 = self.figure2.add_subplot(111)
		if self.ironcounter>0:
			self.axis2 = self.figure2.add_subplot(111) 
			self.axis2.clear()
			graphview.remove(self.canvas2)
			self.builder.get_object("toolbar2").remove(self.toolbar2)
		self.axis2.plot(sortedbkeylist,cvalues, "bo")
		#self.axis2.plot(sortedbkeylist, ironvalues, "b-") #maybe make this within limit later		 
		self.axis2.set_xlabel('Minority Carrier Density (cm$^{-3}$)')   
		self.axis2.set_ylabel('Prefactor, C') #check units are correct   
		self.axis2.ticklabel_format(style='sci',scilimits=(0,0))
		self.axis2.grid(True)
		self.canvas2=FigureCanvasGTKAgg(self.figure2)
		self.canvas2.show()
		graphview.pack_start(self.canvas2, True, True)
		self.toolbar2 = NavigationToolbar(self.canvas2, self.builder.get_object("MainWindow"))
		self.builder.get_object("toolbar2").pack_start(self.toolbar2, False, False)
		self.ironcounter+=1
		self.builder.get_object("resetbtn2").set_sensitive(True)
		self.builder.get_object("saveplotbtn").set_sensitive(True)
		self.builder.get_object("savedatabtn").set_sensitive(False)

		self.builder.get_object("feilabel").set_label("C: ")
		self.builder.get_object("getfebtnlabel").set_label("Get C")
		self.builder.get_object("meanlabel").set_label("Mean of the Prefactor value from\nthe last 10 values: ")	

		cmean= stats.mean(cvalues[-10:-1])
		self.builder.get_object("meanconctxt").set_text("%.4g" % cmean)
		self.curplot="C"

		self.cvalues=cvalues
		self.cvaluedn=sortedbkeylist
		#set default view
		#find ymin
		i=cvalues.index(max(cvalues))+1
		
		while i<len(cvalues):
			if cvalues[i]> (cmean-(abs(cmean)*4)):
				self.axis2.set_ylim(bottom=cvalues[i])
				break
			else:
				i+=1

		self.axis2.set_ylim(top=cmean+(abs(cmean)*4))
		self.builder.get_object("deltangettxt").set_text("")
		self.builder.get_object("fegettxt").set_text("")
		self.builder.get_object("scalebtn2").set_sensitive(True)

	
	def resetbtn2clicked(self, widget):
		if self.curplot=="C":
			self.cfactorbtnclicked(widget)
		elif self.curplot=="Fe":
			self.ironcalcbtnclicked(widget)

			
	def restoredefaultsbtnclicked(self, widget):
		self.builder.get_object("vthermaltxt").set_text("1.1E7")
		self.builder.get_object("sigmanitxt").set_text("1.3E-14")
		self.builder.get_object("p1itxt").set_text("1.28E13")
		self.builder.get_object("sigmapitxt").set_text("7E-17")
		self.builder.get_object("sigmanbtxt").set_text("5E-15")
		self.builder.get_object("n1btxt").set_text("1.22571E15")
		self.builder.get_object("sigmapbtxt").set_text("3E-15")
		
		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]

	def plfileset(self, widget):
		if widget.get_name()=="GtkEntry":
			self.numbify(widget)

		if gtk.Buildable.get_name(widget)=="plafterfile":
			self.adjusted=0

		if self.builder.get_object("plbeforefile").get_filename()!=None and self.builder.get_object("plafterfile").get_filename()!=None and self.builder.get_object("filebeforebtn").get_filename()!=None and self.builder.get_object("fileafterbtn").get_filename()!=None and self.builder.get_object("pldopingtxt").get_text()!=None:
			self.builder.get_object("plcalcbtn").set_sensitive(True)
		else:
			self.builder.get_object("plcalcbtn").set_sensitive(False)

		if gtk.Buildable.get_name(widget)=="plbeforefile":
			#display dialog asking for generation level
			self.builder.get_object("pl1genleveldialog").show()
		
	def plcalcbtnclicked(self, widget):
		#try:

		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]

		plbeforefile = open(self.builder.get_object("plbeforefile").get_filename(), "r")
		beforeraw = loadtxt(plbeforefile, dtype=float)
		self.taubefore = 1E6*beforeraw
		b = 1E6*beforeraw
		plbeforefile.close()
	       	#savetxt("beforefilesave.txt", self.taubefore, fmt="%12.6G")
		if self.adjusted==0:
			plafterfile = open(self.builder.get_object("plafterfile").get_filename(), "r")
			afterraw=loadtxt(plafterfile, dtype=float)
			self.tauafter=1E6*afterraw
			a=1E6*afterraw
			plafterfile.close()
		else:
			a=np.copy(self.tauafter)
		#accounts for some files using already microseconds
		if self.taubefore.mean()>1000:
			self.taubefore=self.taubefore/1E6
		if self.tauafter.mean()>1000:
			self.tauafter=self.tauafter/1E6
		# 1E6 factor is necessary to put in microseconds

		#Need injection level
		injectionlevel=float(self.builder.get_object("injectionleveltxt").get_text())
		C=concentration.calcPrefactor(self.constants, float(self.builder.get_object("pldopingtxt").get_text()), injectionlevel)
		self.ironconcmatrix = abs(C*(1/self.taubefore - 1/self.tauafter))
		#HACK TO MAKE NEGATIVE VALUES POSITIVE
		e=C*(1/self.taubefore - 1/self.tauafter)
		#throws errors on some values may need to fix for py2exe
		#scale stuff here - must chop off extreme values, copy alex's way
		#numpy arrays pass by reference, they do NOT create a copy also only allows one instance open?
		a1 = a[500]
		b1 = b[500]
		e1 = e[500]
		a1.sort()
		b1.sort()
		e1.sort()
		self.taubmin = b1[int(0.01*len(b))]
		self.taubmax = b1[int(0.99*len(b))]
		self.tauamin = a1[int(0.01*len(a))]
		self.tauamax = a1[int(0.99*len(a))]
		self.ironmin = e1[int(0.04*len(e))]
		#if self.ironmin<0:
		#	self.ironmin=0
		#see about making colors different when off scale
		self.ironmax = e1[int(0.97*len(e))]

		self.builder.get_object("lifemintxt").set_text("%.4g" %self.taubmin)
		self.builder.get_object("lifemaxtxt").set_text("%.4g" %self.taubmax)
		self.builder.get_object("femintxt").set_text("%.4g" %self.ironmin)
		self.builder.get_object("femaxtxt").set_text("%.4g" %self.ironmax)
		self.plplotallbtnclicked(widget)
		#random numbers chosen away from edges to avoid infinities
		if self.ironconcmatrix[547:555].mean()<=0:
			self.builder.get_object("negiron").show()


		#except:
			#self.errorshow(widget)

		if self.adjusted==0:
		#cross-correlation - ignored until can fix Mirco's program
#			self.oldcc=((self.tauafter-self.tauafter.mean()) * (self.taubefore - self.taubefore.mean())).mean() / (self.taubefore.std() * self.tauafter.std())
			self.oldcc=bildreg.correlation(self.taubefore[100:-100,100:-100], self.tauafter[100:-100,100:-100])
			#if cc<0.95:
			#	self.builder.get_object("correlationwarning").set_property("secondary-text", "There was a significant difference between the loaded PL images, this means there may have been a change in the position of the sample between the taking of the images.\nCross-correlation value: %.4g (anything below 0.95 is considered anomalous)" % cc)
			#	self.builder.get_object("correlationwarning").show()

			self.builder.get_object("correlationlabel").set_label("The Cross-Correlation value for the images is %.4g\n <b>Should image matching be attempted?</b>\n(note this only guarantees an improvement below 0.85, at high correlations it may produce a worse image!)\n" % self.oldcc) 
			self.builder.get_object("imagematchingwindow").show()



	def attemptcorrectionbtnclicked(self, widget):
		trafo = bildreg.transformation_berechnen(self.taubefore, self.tauafter, maxdelta=34, korrelationsgroesse=128, N=5)
		T_T = np.ndarray(self.tauafter.shape, dtype=np.float)
		bildreg.transform(self.tauafter,trafo,T_T)
		self.tauafter=T_T
		self.adjusted=1
		self.builder.get_object("imagematchingwindow").hide()
		self.plcalcbtnclicked(widget)
		self.newcc=bildreg.correlation(self.taubefore[100:-100,100:-100], self.tauafter[100:-100,100:-100])
#((self.tauafter-self.tauafter.mean()) * (self.taubefore - self.taubefore.mean())).mean() / (self.taubefore.std() * self.tauafter.std())
		self.builder.get_object("matchingcomplete").set_property("secondary-text", "The lifetime map after illumination was translated on to the lifetime map before illumination. The new cross-correlation value is %.4g, compared to the old one of %.4g" % (self.newcc, self.oldcc))
		self.builder.get_object("matchingcomplete").show()
		
	def donothingbtnclicked(self, widget):
		self.builder.get_object("imagematchingwindow").hide()
		
	def plplotallbtnclicked(self, widget):

		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]

		if self.plcounter==0:
			self.figure3 = Figure(figsize=(6,4), dpi=72)  
			self.figure4=Figure(figsize=(6,4), dpi=72)
		if self.plcounter>0:
			self.axis3.clear()
			self.axis4.clear()
			self.axis5.clear()
			self.figure3.clear()
			self.figure4.clear()
			self.builder.get_object("plgraphs1").remove(self.canvas3)
			self.builder.get_object("plgraphs2").remove(self.canvas4)
			l_f = LogFormatter(10, labelOnlyBase=False)
		if self.lifemaptype=="logarithmic":
			self.axis3 = self.figure3.add_subplot(211, aspect='equal')
			imagebefore=self.axis3.imshow(self.taubefore, cmap=self.cmap, norm=LogNorm(vmin=self.taubmin, vmax=self.taubmax))
			cbar1 = self.figure3.colorbar(imagebefore, format = l_f, extend='both') 
			self.axis3.set_title("Lifetime before illumination")
			self.axis4=self.figure3.add_subplot(212, aspect='equal')
			imageafter=self.axis4.imshow(self.tauafter, cmap=self.cmap, norm=LogNorm(vmin=self.taubmin, vmax=self.taubmax))
			cbar2=self.figure3.colorbar(imageafter, format=l_f, extend='both')

		elif self.lifemaptype=="linear":
			self.axis3 = self.figure3.add_subplot(211, aspect='equal')
			imagebefore=self.axis3.imshow(self.taubefore, cmap=self.cmap,vmin=self.taubmin, vmax=self.taubmax)
			cbar1=self.figure3.colorbar(imagebefore, orientation="vertical", extend='both')
			self.axis3.set_title("Lifetime before illumination")
			self.axis4=self.figure3.add_subplot(212, aspect='equal')
			imageafter=self.axis4.imshow(self.tauafter, cmap=self.cmap, vmin=self.taubmin, vmax=self.taubmax)
			cbar2=self.figure3.colorbar(imageafter, extend='both')

		cbar1.set_label("Lifetime ($\mu$s)", rotation='vertical')
		cbar2.ax.set_ylabel("Lifetime ($\mu$s)", rotation='vertical', verticalalignment='center') #rotation doesnt work with mathtex
		self.axis4.set_title("Lifetime after illumination")
		self.canvas3=FigureCanvasGTKAgg(self.figure3)
		self.canvas3.show()

		self.axis5=self.figure4.add_subplot(111, aspect='equal')
		if self.ironmaptype=="linear":
			imageiron=self.axis5.imshow(self.ironconcmatrix, cmap=self.cmap, vmin=self.ironmin, vmax=self.ironmax)
			cbar3=self.figure4.colorbar(imageiron, fraction=0.045, extend='both')
		elif self.ironmaptype=="logarithmic":
			imageiron=self.axis5.imshow(self.ironconcmatrix, cmap=self.cmap, norm=LogNorm(vmin=self.ironmin, vmax=self.ironmax))
			cbar3=self.figure4.colorbar(imageiron, format=l_f, fraction=0.045, extend='both')

		cbar3.set_label(r'[Fe$_{i}$]  (cm$^{-3})$')
		self.axis5.set_title("Interstitial Iron concentration")
		self.canvas4=FigureCanvasGTKAgg(self.figure4)
		self.canvas4.show()
		
		self.axis3.set_xlim(left=0, right=1000)
		self.axis3.set_ylim(top=0, bottom=1000)
		self.axis4.set_xlim(left=0, right=1000)
		self.axis4.set_ylim(top=0, bottom=1000)
		self.axis5.set_xlim(left=0, right=1000)
		self.axis5.set_ylim(top=0, bottom=1000)
		
		self.builder.get_object("plgraphs1").pack_start(self.canvas3, True, True)
		self.builder.get_object("plgraphs2").pack_start(self.canvas4, True, True)
		self.builder.get_object("pltoolbar").set_sensitive(True)
		self.builder.get_object("editcolorbarbtn").set_sensitive(True)
		self.plcounter+=1

		self.canvas4.mpl_connect('motion_notify_event', self.graphscroll)
		self.canvas3.mpl_connect('motion_notify_event', self.graphscroll)

	def zoominbtnclicked(self, widget):
		xlims=self.axis5.get_xlim()
		ylims=self.axis5.get_ylim()
		xchange=abs(xlims[1]-xlims[0])*0.1
		ychange=abs(ylims[1]-ylims[0])*0.1
		self.axis3.set_xlim(left=xlims[0]+xchange, right=xlims[1]-xchange)
		self.axis3.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]-ychange)
		self.axis4.set_xlim(left=xlims[0]+xchange, right=xlims[1]-xchange)
		self.axis4.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]-ychange)
		self.axis5.set_xlim(left=xlims[0]+xchange, right=xlims[1]-xchange)
		self.axis5.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]-ychange)
		self.builder.get_object("plgraphs1").remove(self.canvas3)
		self.builder.get_object("plgraphs2").remove(self.canvas4)
		self.builder.get_object("plgraphs1").pack_start(self.canvas3, True, True)
		self.builder.get_object("plgraphs2").pack_start(self.canvas4, True, True)
		
	def zoomoutbtnclicked(self, widget):
		xlims=self.axis5.get_xlim()
		ylims=self.axis5.get_ylim()
		xchange=abs(xlims[1]-xlims[0])*0.1
		xchange=xchange/0.8
		ychange=abs(ylims[1]-ylims[0])*0.1
		ychange=ychange/0.8
		self.axis3.set_xlim(left=xlims[0]-xchange, right=xlims[1]+xchange)
		self.axis3.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]+ychange)
		self.axis4.set_xlim(left=xlims[0]-xchange, right=xlims[1]+xchange)
		self.axis4.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]+ychange)
		self.axis5.set_xlim(left=xlims[0]-xchange, right=xlims[1]+xchange)
		self.axis5.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]+ychange)
		self.builder.get_object("plgraphs1").remove(self.canvas3)
		self.builder.get_object("plgraphs2").remove(self.canvas4)
		self.builder.get_object("plgraphs1").pack_start(self.canvas3, True, True)
		self.builder.get_object("plgraphs2").pack_start(self.canvas4, True, True)

       	def panleftbtnclicked(self, widget):
		xlims=self.axis5.get_xlim()
		xchange=abs(xlims[1]-xlims[0])*0.1
		self.axis3.set_xlim(left=xlims[0]-xchange, right=xlims[1]-xchange)
		self.axis4.set_xlim(left=xlims[0]-xchange, right=xlims[1]-xchange)
		self.axis5.set_xlim(left=xlims[0]-xchange, right=xlims[1]-xchange)
		self.builder.get_object("plgraphs1").remove(self.canvas3)
		self.builder.get_object("plgraphs2").remove(self.canvas4)
		self.builder.get_object("plgraphs1").pack_start(self.canvas3, True, True)
		self.builder.get_object("plgraphs2").pack_start(self.canvas4, True, True)

       	def panrightbtnclicked(self, widget):
		xlims=self.axis5.get_xlim()
		xchange=abs(xlims[1]-xlims[0])*0.1
		self.axis3.set_xlim(left=xlims[0]+xchange, right=xlims[1]+xchange)
		self.axis4.set_xlim(left=xlims[0]+xchange, right=xlims[1]+xchange)
		self.axis5.set_xlim(left=xlims[0]+xchange, right=xlims[1]+xchange)
		self.builder.get_object("plgraphs1").remove(self.canvas3)
		self.builder.get_object("plgraphs2").remove(self.canvas4)
		self.builder.get_object("plgraphs1").pack_start(self.canvas3, True, True)
		self.builder.get_object("plgraphs2").pack_start(self.canvas4, True, True)

	def panupbtnclicked(self, widget):
		ylims=self.axis5.get_ylim()
		ychange=abs(ylims[1]-ylims[0])*0.1
		self.axis3.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]-ychange)
		self.axis4.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]-ychange)
		self.axis5.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]-ychange)
		self.builder.get_object("plgraphs1").remove(self.canvas3)
		self.builder.get_object("plgraphs2").remove(self.canvas4)
		self.builder.get_object("plgraphs1").pack_start(self.canvas3, True, True)
		self.builder.get_object("plgraphs2").pack_start(self.canvas4, True, True)

	def pandownbtnclicked(self, widget):
		ylims=self.axis5.get_ylim()
		ychange=abs(ylims[1]-ylims[0])*0.1
		self.axis3.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]+ychange)
		self.axis4.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]+ychange)
		self.axis5.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]+ychange)
		self.builder.get_object("plgraphs1").remove(self.canvas3)
		self.builder.get_object("plgraphs2").remove(self.canvas4)
		self.builder.get_object("plgraphs1").pack_start(self.canvas3, True, True)
		self.builder.get_object("plgraphs2").pack_start(self.canvas4, True, True)
	
	def savefemapbtnclicked(self, widget):
		self.builder.get_object("whichsavewindow").show()

	def whichsavefebtnclicked(self, widget):
		self.cursavepl="femap"
		self.builder.get_object("whichsavewindow").hide()
		self.builder.get_object("plplotfilesavedialog").set_property("title", "Save interstitial Iron concentration map to PNG")
		self.builder.get_object("plplotfilesavedialog").show()

	def whichsavelifebtnclicked(self, widget):
		self.cursavepl="lifemap"
		self.builder.get_object("whichsavewindow").hide()
		self.builder.get_object("plplotfilesavedialog").set_property("title", "Save lifetime maps to PNG")
		self.builder.get_object("plplotfilesavedialog").show()


	def whichsavecancelbtnclicked(self, widget):
		self.builder.get_object("whichsavewindow").hide()		

	def plplotdialogsavebtnclicked(self, widget):
		#atm only does iron map
		self.issaving=True
		self.cursave="pl"
		filename=self.builder.get_object("plplotfilesavedialog").get_filename()
		if filename[-4:len(filename)]!=".png" and filename[-4:len(filename)]!=".PNG":
			filename += ".png"
		self.currentfilename=filename
		if os.path.exists(filename)==True:
			self.builder.get_object("overwritewarning").show()
		else:
			self.saveplmap()

			
	
	def plplotdialogcancelbtnclicked(self, widget):
		self.builder.get_object("plplotfilesavedialog").hide()
		
	def overwriteresponseclicked(self, widget, response):
		#yes -8 no -9
		self.issaving=True
		self.builder.get_object("overwritewarning").hide()
		if response==-8:
			if self.cursave=="pl":
				self.saveplmap()
			elif self.cursave=="plotdialogsave":
				self.figure2.savefig(self.currentfilename, format="png")
				self.builder.get_object("plotfilesavedialog").hide()	
			elif self.cursave=="feplotdata":
				self.datatxtsave()
			elif self.cursave=="femapdata":
				savetxt(self.currentfilename, self.ironconcmatrix, fmt="%12.6G")
			elif self.cursave=="adjusteddata":
				savetxt(self.currentfilename, self.tauafter, fmt="%12.6G")
	def saveplmap(self):
		self.builder.get_object("plplotfilesavedialog").hide()
		#add error handling
		#check if filename has extension, if not append PNG - maybe ask?
		l_f = LogFormatter(10, labelOnlyBase=False)
		if self.cursavepl=="femap":
			xlims=self.axis5.get_xlim()
			ylims=self.axis5.get_ylim()
			if self.issaving==True:
				self.figure4.savefig(self.currentfilename, format="png")
			self.axis5.clear()
			self.figure4.clear()
			self.builder.get_object("plgraphs2").remove(self.canvas4)
			self.axis5=self.figure4.add_subplot(111, aspect='equal')

			if self.ironmaptype=="linear":
				imageiron=self.axis5.imshow(self.ironconcmatrix, cmap=self.cmap, vmin=self.ironmin, vmax=self.ironmax)
				cbar3=self.figure4.colorbar(imageiron, fraction=0.045, extend='both')
			elif self.ironmaptype=="logarithmic":
				imageiron=self.axis5.imshow(self.ironconcmatrix, cmap=self.cmap, norm=LogNorm(vmin=self.ironmin, vmax=self.ironmax))
				cbar3=self.figure4.colorbar(imageiron, fraction=0.045, format=l_f, extend='both')

			cbar3.set_label(r'[Fe$_{i}$]  (cm$^{-3})$')
			self.axis5.set_title("Interstitial Iron concentration")
			self.canvas4=FigureCanvasGTKAgg(self.figure4)
			self.canvas4.show()

			self.axis5.set_xlim(xlims)
			self.axis5.set_ylim(ylims)

			self.builder.get_object("plgraphs2").pack_start(self.canvas4, True, True)
			self.canvas4.mpl_connect('motion_notify_event', self.graphscroll)

		elif self.cursavepl=="lifemap":
			xlims=self.axis5.get_xlim()
			ylims=self.axis5.get_ylim()
			if self.issaving==True:
				self.figure3.savefig(self.currentfilename, format="png")

			self.axis3.clear()
			self.axis4.clear()
			self.figure3.clear()
			self.builder.get_object("plgraphs1").remove(self.canvas3)
			
			if self.lifemaptype=="logarithmic":
				self.axis3 = self.figure3.add_subplot(211, aspect='equal')
				imagebefore=self.axis3.imshow(self.taubefore, cmap=self.cmap, norm=LogNorm(vmin=self.taubmin, vmax=self.taubmax))
				cbar1 = self.figure3.colorbar(imagebefore, format = l_f, extend='both') 
				self.axis3.set_title("Lifetime before illumination")
				self.axis4=self.figure3.add_subplot(212, aspect='equal')
				imageafter=self.axis4.imshow(self.tauafter, cmap=self.cmap, norm=LogNorm(vmin=self.taubmin, vmax=self.taubmax))
				cbar2=self.figure3.colorbar(imageafter, format=l_f, extend='both')

			elif self.lifemaptype=="linear":
				self.axis3 = self.figure3.add_subplot(211, aspect='equal')
				imagebefore=self.axis3.imshow(self.taubefore, cmap=self.cmap,vmin=self.taubmin, vmax=self.taubmax)
				cbar1=self.figure3.colorbar(imagebefore, orientation="vertical", extend='both')
				self.axis3.set_title("Lifetime before illumination")
				self.axis4=self.figure3.add_subplot(212, aspect='equal')
				imageafter=self.axis4.imshow(self.tauafter, cmap=self.cmap, vmin=self.taubmin, vmax=self.taubmax)
				cbar2=self.figure3.colorbar(imageafter, extend='both')


			cbar1.set_label("Lifetime ($\mu$s)", rotation='vertical', verticalalignment='center')
			cbar2.set_label("Lifetime ($\mu$s)")
			self.axis4.set_title("Lifetime after illumination")
			self.canvas3=FigureCanvasGTKAgg(self.figure3)
			self.canvas3.show()

			self.axis3.set_xlim(xlims)
			self.axis3.set_ylim(ylims)
			self.axis4.set_xlim(xlims)
			self.axis4.set_ylim(ylims)
			self.canvas3.mpl_connect('motion_notify_event', self.graphscroll)

			self.builder.get_object("plgraphs1").pack_start(self.canvas3, True, True)

	def pl1genlevelokbtnclicked(self, widget):
		try:
			self.builder.get_object("pl1genleveldialog").hide()
			genlevelaim=float(self.builder.get_object("pl1genleveltxt").get_text())
			filename1=self.builder.get_object("filebeforebtn").get_filename()
			closestlevel=0
			closesti=0
			wb1 = load_workbook(filename1)
			ws1 = wb1.get_sheet_by_name("Calc")
			i=15
			for row in ws1.range('K15:K131'):
				for cell in row:
					curdiff=abs(genlevelaim-float(cell.value))
					if curdiff<abs(genlevelaim-closestlevel):
						closestlevel=float(cell.value)
						closesti=i

					i+=1

			injectionlevelaim=float(ws1.cell('S'+str(closesti)).value)
			self.injectionlevel=injectionlevelaim
			#get wanted generation level for second file
			closestlevel=0
			closesti=0
			filename2=self.builder.get_object("fileafterbtn").get_filename()
			wb2 = load_workbook(filename2)
			ws2 = wb2.get_sheet_by_name("Calc")
			i=15
			for row in ws2.range('S15:S131'):
				for cell in row:
					curdiff=abs(injectionlevelaim-float(cell.value))
					if curdiff<abs(injectionlevelaim-closestlevel):
						closestlevel=float(cell.value)
						closesti=i

					i+=1

			wantedilevel=float(ws2.cell('K'+str(closesti)).value)
			#print float(ws2.cell('S'+str(closesti)).value)
			self.builder.get_object("plgenleveltxt").set_text("%.4g" % wantedilevel)
			self.builder.get_object("injectionleveltxt").set_text("%.4g" % injectionlevelaim)
			self.builder.get_object("plafterfile").set_property("title", "Open PL data after illumination, G~%.3g" % wantedilevel)

			self.builder.get_object("plafterfile").set_sensitive(True)
			
		except:
			#change this to a proper error message later
			self.errorshow(self)
			return 1

	def pl1genlevelcancelbtnclicked(self,widget):
		self.builder.get_object("pl1genleveldialog").hide()

	def getfebtnclicked(self, widget):
		
		if self.curplot=="Fe":
			#limit=self.irondata[1].index(max(self.irondata[1]))
			ironvalues=self.irondata[1]
			deltanvalues=self.irondata[0]
			interpf=interp1d(deltanvalues, ironvalues)
			try: 
				interpval=interpf(float(self.builder.get_object("deltangettxt").get_text()))
				#interpval=interp(float(self.builder.get_object("deltangettxt").get_text()), deltanvalues, ironvalues, left=-99.0, right=-99.0)
				#add error handling code
				self.builder.get_object("fegettxt").set_text("%.3e" % interpval)


			except:
				self.builder.get_object("interperrorwindow").show()
		elif self.curplot=="C":
			deltanvalues=self.cvaluedn
			interpf=interp1d(deltanvalues, self.cvalues)
			try: 
				interpval=interpf(float(self.builder.get_object("deltangettxt").get_text()))
				#interpval=interp(float(self.builder.get_object("deltangettxt").get_text()), deltanvalues, self.cvalues, left=-99.0, right=-99.0)
				self.builder.get_object("fegettxt").set_text("%.3e" % interpval)


			except:
				self.builder.get_object("interperrorwindow").show()
			
	def editcolorbarbtnclicked(self, widget):
	        self.builder.get_object("lifemintxt").set_text("%.4g" %self.taubmin)
		self.builder.get_object("lifemaxtxt").set_text("%.4g" %self.taubmax)
		self.builder.get_object("femintxt").set_text("%.4g" %self.ironmin)
		self.builder.get_object("femaxtxt").set_text("%.4g" %self.ironmax)
		if self.lifemaptype=="linear":
			self.linbutton.set_active(True)
		if self.lifemaptype=="logarithmic":
			self.logbutton.set_active(True)
		if self.ironmaptype=="linear":
			self.felinbutton.set_active(True)
		if self.ironmaptype=="logarithmic":
			self.felogbutton.set_active(True)
		if self.cmap==cm.gray:
			self.combobox.set_active(0)
		elif self.cmap==cm.jet:
			self.combobox.set_active(1)
		elif self.cmap==cm.hot:
			self.combobox.set_active(2)
		self.builder.get_object("editcolorbarwindow").show()

	def editcolorbarcancelbtnclicked(self,widget):
		self.builder.get_object("editcolorbarwindow").hide()

	def editcolorbarokbtnclicked(self,widget):
		self.builder.get_object("editcolorbarwindow").hide()
	        self.taubmin=float(self.builder.get_object("lifemintxt").get_text())
	        self.taubmax=float(self.builder.get_object("lifemaxtxt").get_text())
	        self.ironmin=float(self.builder.get_object("femintxt").get_text())
	        self.ironmax=float(self.builder.get_object("femaxtxt").get_text())
		if self.combobox.get_active()==0:
			self.cmap=cm.gray
		elif self.combobox.get_active()==1:
			self.cmap=cm.jet
		elif self.combobox.get_active()==2:
			self.cmap=cm.hot
		self.ironmaptype=self.tempironmaptype
		self.lifemaptype=self.templifemaptype
		self.issaving=False
		self.cursavepl="femap"
		self.saveplmap()
		self.cursavepl="lifemap"
		self.saveplmap()

	def saveirondatabtnclicked(self, widget):
		self.datasave="map"
		self.builder.get_object("whichsavewindow").hide()
		self.builder.get_object("datafilesavedialog").show()

	def buttontoggle(self, widget, data=None):
		if widget.get_active()==True:
			self.templifemaptype=data

	def febuttontoggle(self, widget, data=None):
		if widget.get_active()==True:
			self.tempironmaptype=data

	def saveadjusteddatabtnclicked(self, widget):
		self.datasave="adjusteddata"
		self.builder.get_object("datafilesavedialog").show()
		

	def graphscroll(self, event):
		if event.x!=None and event.y!=None and event.xdata!=None and event.ydata!=None:

			if self.curid!=None:
				self.builder.get_object("statusbar").remove_message(self.plconid, self.curid)
				
			self.curid=self.builder.get_object("statusbar").push(self.plconid, 'x=%d, y=%d, taub=%.3g, taua=%.3g, iron=%.3g'%(int(round(event.xdata)), int(round(event.ydata)), self.taubefore[int(round(event.ydata))][int(round(event.xdata))], self.tauafter[int(round(event.ydata))][int(round(event.xdata))], self.ironconcmatrix[int(round(event.ydata))][int(round(event.xdata))]))

			print 'x=%d, y=%d, taub=%.3g, taua=%.3g, iron=%.3g'%(int(round(event.xdata)), int(round(event.ydata)), self.taubefore[int(round(event.ydata))][int(round(event.xdata))], self.tauafter[int(round(event.ydata))][int(round(event.xdata))], self.ironconcmatrix[int(round(event.ydata))][int(round(event.xdata))])


if __name__ == "__main__":
	app = MyApp()
	gtk.main()
