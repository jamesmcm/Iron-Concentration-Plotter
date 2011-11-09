import matplotlib 
matplotlib.use('GTK')  
import pygtk
pygtk.require("2.0")
import gtk
import excel
from numpy import savetxt, loadtxt
import numpy as np
from scipy.interpolate import interp1d
import os
from matplotlib.pyplot import plot, savefig, rc
from matplotlib.colorbar import Colorbar, ColorbarBase
from matplotlib.figure import Figure   
from matplotlib.ticker import LogLocator, LogFormatter 
import matplotlib.cm as cm
from math import pow
from matplotlib.axes import Subplot   
from matplotlib.backends.backend_gtkagg import NavigationToolbar2GTKAgg as NavigationToolbar
from matplotlib.backends.backend_gtkagg import FigureCanvasGTKAgg  
import fitting
import doping
import bildregistrierung_ng as bildreg
from openpyxl.reader.excel import load_workbook
import concentration
from matplotlib.colors import LogNorm
#from scipy import stats
import scipy.interpolate

import warnings
warnings.simplefilter('ignore')

#matplotlib.rc('text', usetex=True) doesn't work

class MyApp(object):       
	def __init__(self):
	    self.builder = gtk.Builder()
	    self.builder.add_from_file("myapp.xml")
	    self.builder.connect_signals({ "on_MainWindow_destroy" : gtk.main_quit, "on_menuquit_activate" : gtk.main_quit, "on_filebeforebtn_file_set" : self.plotgraph1, "on_ErrorWindow_close" : self.errorclose, "scaletxtedit":self.scaletxtedit, "scalebtnclicked":self.scalebtnclicked, "resetbtnclicked":self.plotgraph1, "resistivityresponse" : self.errorclose, "ironcalcbtnclicked" : self.ironcalcbtnclicked, "scalebtn2clicked":self.scalebtn2clicked, "resetbtn2clicked":self.resetbtn2clicked, "plotdialogcancelbtnclicked":self.plotdialogcancelbtnclicked, "plotdialogsavebtnclicked":self.plotdialogsavebtnclicked, "savedatabtnclicked":self.savedatabtnclicked, "datadialogsavebtnclicked":self.datadialogsavebtnclicked, "datadialogcancelbtnclicked":self.datadialogcancelbtnclicked, "cfactorbtnclicked":self.cfactorbtnclicked, "plotflippeddialogclose":self.errorclose, "restoredefaultsbtnclicked":self.restoredefaultsbtnclicked, "plcalcbtnclicked":self.plcalcbtnclicked, "plfileset":self.plfileset, "plplotallbtnclicked":self.plplotallbtnclicked, "zoominbtnclicked":self.zoominbtnclicked, "zoomoutbtnclicked":self.zoomoutbtnclicked, "panleftbtnclicked":self.panleftbtnclicked, "panrightbtnclicked":self.panrightbtnclicked, "panupbtnclicked":self.panupbtnclicked, "pandownbtnclicked":self.pandownbtnclicked, "plplotdialogsavebtnclicked": self.plplotdialogsavebtnclicked, "plplotdialogcancelbtnclicked":self.plplotdialogcancelbtnclicked, "savefemapbtnclicked":self.savefemapbtnclicked, "overwriteresponseclicked":self.overwriteresponseclicked, "numbify":self.numbify, "pl2genlevelokbtnclicked":self.pl2genlevelokbtnclicked, "pl2genlevelcancelbtnclicked":self.pl2genlevelcancelbtnclicked, "getfebtnclicked":self.getfebtnclicked, "whichsavefebtnclicked":self.whichsavefebtnclicked, "whichsavelifebtnclicked":self.whichsavelifebtnclicked, "whichsavecancelbtnclicked":self.whichsavecancelbtnclicked, "editcolorbarbtnclicked":self.editcolorbarbtnclicked, "editcolorbarokbtnclicked":self.editcolorbarokbtnclicked, "editcolorbarcancelbtnclicked":self.editcolorbarcancelbtnclicked, "saveirondatabtnclicked":self.saveirondatabtnclicked, "recalccopbtnclicked":self.recalccopbtnclicked, "restoredopingbtnclicked":self.restoredopingbtnclicked, "attemptcorrectionbtnclicked":self.attemptcorrectionbtnclicked, "donothingbtnclicked":self.donothingbtnclicked, "saveadjusteddatabtnclicked":self.saveadjusteddatabtnclicked, "calcdopefromresbtnclicked":self.calcdopefromresistivity, "resbeforevalbtnclicked":self.resbeforevalbtnclicked, "resaftervalbtnclicked":self.resaftervalbtnclicked, "resmyvalbtnclicked":self.resmyvalbtnclicked,"showironvaluesbtnclicked":self.showironvaluesbtnclicked,"ironviewclosebtnclicked":self.ironviewclosebtnclicked, "editfitokbtnclicked":self.editfitokbtnclicked, "fitdefaultsbtnclicked":self.fitdefaultsbtnclicked, "editfitcancelbtnclicked":self.editfitcancelbtnclicked, "changefitparamsbtnclicked":self.changefitparamsbtnclicked, "calculatemeanrangeclicked":self.calculatemeanrangeclicked, "compareplmapsbtnclicked":self.compareplmapsbtnclicked, "closecomparebtnclicked":self.closecomparebtnclicked, "pandowncomparebtnclicked":self.pandowncomparebtnclicked, "panupcomparebtnclicked":self.panupcomparebtnclicked, "panleftcomparebtnclicked":self.panleftcomparebtnclicked, "panrightcomparebtnclicked":self.panrightcomparebtnclicked, "zoomincomparebtnclicked":self.zoomincomparebtnclicked, "zoomoutcomparebtnclicked":self.zoomoutcomparebtnclicked, "refreshcomparebtnclicked":self.refreshcomparebtnclicked})
	    self.window = self.builder.get_object("MainWindow")
	    #self.window.fullscreen()
	    filter1 = gtk.FileFilter()
	    filter1.set_name("XLSM")
	    filter1.add_pattern("*.xlsm")
	    filter2 = gtk.FileFilter()
	    filter2.set_name("TXT")
	    filter2.add_pattern("*.txt")
	    filterplot = gtk.FileFilter()
	    filterplot.set_name("PNG")
	    filterplot.add_pattern("*.png")
	    self.builder.get_object("plotfilesavedialog").add_filter(filterplot)
	    self.builder.get_object("plplotfilesavedialog").add_filter(filterplot)
	    self.builder.get_object("plbeforefile").add_filter(filter2)
	    self.builder.get_object("plafterfile").add_filter(filter2)
	    self.builder.get_object("datafilesavedialog").add_filter(filter2)
	    
	    self.beforefileselect=self.builder.get_object("filebeforebtn")
	    self.beforefileselect.add_filter(filter1)
	    self.afterfileselect=self.builder.get_object("fileafterbtn")
	    self.afterfileselect.add_filter(filter1)
	    self.graph1counter=0
	    self.ironcounter=0
	    self.adjusted=0
	    self.plcounter=0
	    self.cursave=""
	    self.curid=None
	    self.curid2=None	    
	    self.fitparams=[]
	    #taup, taun, n1, p1, NA
	    self.startguessb=np.array([1.72e-3,1e-2,1.22E2,9e1,2e15],np.float64)
	    self.startguessa=np.array([1.72e-3,1e-2,1.22E2,9e1,2e15],np.float64)
	    self.defaultstartguess=np.array([1.72e-3,1e-2,1.22E2,9e1,2e15],np.float64)

	    self.combobox = gtk.combo_box_new_text()
	    self.builder.get_object("combospace").add(self.combobox)
	    self.builder.get_object("combospace").show_all()
	    self.combobox.append_text('Gray')
	    self.combobox.append_text('Jet')
	    self.combobox.append_text('Hot')
	    self.combobox.set_active(0)

	    self.linbutton = gtk.RadioButton(None, "Linear")
	    self.linbutton.connect("toggled", self.buttontoggle, "linear")
	    self.builder.get_object("rbutton1vbox").pack_start(self.linbutton, True, True, 0)
	    self.linbutton.show()
	    self.logbutton = gtk.RadioButton(self.linbutton, "Base-10 Logarithmic")
	    self.logbutton.connect("toggled", self.buttontoggle, "logarithmic")
	    self.builder.get_object("rbutton2vbox").pack_start(self.logbutton, True, True, 0)
	    self.logbutton.show()
	    self.linbutton.set_active(True)
	    #self.logbutton.set_active(False)
	    self.lifemaptype="linear"
	    self.templifemaptype="linear"

	    #iron plot type: individual - individual C factor values, genlevel - Cfactor based on gen level, mean - Use mean of last Cfactor values, indgen - individual C values based on gen levels and tau values
	    self.ironplottype="indgen"
	    self.ironplottypebox = gtk.combo_box_new_text()
	    self.builder.get_object("ironplottypespace").add(self.ironplottypebox)
	    self.builder.get_object("ironplottypespace").show_all()
	    self.ironplottypebox.append_text('Individual C values from Generation Levels and Tau values')	   
	    self.ironplottypebox.append_text('Individual C values from QSSPC Fits')
	    self.ironplottypebox.append_text('C value from Generation Level')
	    self.ironplottypebox.append_text('Use mean C value in given range')
	    self.ironplottypebox.set_active(0)

	    self.fitusebox = gtk.combo_box_new_text()
	    self.builder.get_object("usefitcombospace").add(self.fitusebox)
	    self.builder.get_object("usefitcombospace").show_all()
	    self.fitusebox.append_text('Use fits')
	    self.fitusebox.append_text('Use interpolation')
	    self.fitusebox.connect("changed", self.fituseboxchanged)
	    self.fitusebox.set_active(0)

	    self.delimbox = gtk.combo_box_new_text()
	    self.builder.get_object("delimvbox").add(self.delimbox)
	    self.builder.get_object("delimvbox").show_all()
	    self.delimbox.append_text('Periods .')
	    self.delimbox.append_text('Commas ,')
	    self.delimbox.set_active(0)


	    #iron radio buttons
	    self.felinbutton = gtk.RadioButton(None, "Linear")
	    self.felinbutton.connect("toggled", self.febuttontoggle, "linear")
	    self.builder.get_object("ironradio1vbox").pack_start(self.felinbutton, True, True, 0)
	    self.felinbutton.show()
	    self.felogbutton = gtk.RadioButton(self.felinbutton, "Base-10 Logarithmic")
	    self.felogbutton.connect("toggled", self.febuttontoggle, "logarithmic")
	    self.builder.get_object("ironradio2vbox").pack_start(self.felogbutton, True, True, 0)
	    self.felogbutton.show()
	    self.felinbutton.set_active(True)
	    self.ironmaptype="linear"
	    self.tempironmaptype="linear"
	    self.plconid=self.builder.get_object("statusbar").get_context_id("graphs")
	    self.plconid2=self.builder.get_object("comparestatusbar").get_context_id("maps")	    
	    self.constants=[1.1E7,1.3E-14,1.28E13,7E-17,5E-15,1.22571E15,3E-15]
	    self.window.show()
	    cm.gray.set_under('r') #add these to colorbar somehow?
	    #cm.gray.set_over('b')
	    #cm.gray.set_bad('g')
	    cm.jet.set_under('w') 
	   #cm.jet.set_over('k')
	    cm.hot.set_under('b')
	    #cm.hot.set_over('g')
	    self.cmap=cm.gray
	    self.canfit=1
	    self.afterfitplot=[]
	    self.beforefitplot=[]
	    self.bfitrmin=0
	    self.afitrmin=0
	    self.bfitrmax=0
	    self.afitrmax=0
	    self.befile=0
	    self.affile=0
	    self.beforeplotcounter=0
	    self.afterplotcounter=0
#	    self.ddnlimit=1e14
	    self.ddnlimit=1e14	    
	    self.ironplmapcounter=0
	    self.meanmode="Fe" #or C
	    self.fitplotlimit=0
	    self.compcounter=0

	def fituseboxchanged(self, widget):
		self.plotgraph1(widget)
	
	def plotgraph1(self, widget):
		befile=self.builder.get_object("filebeforebtn")
		affile=self.builder.get_object("fileafterbtn")
		graphview = self.builder.get_object("graph1")  
		#remember the order of the constants
		#vthermal sigmani p1i sigmapi sigmanb n1b sigmapb
		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]
		if self.graph1counter==0:
			self.figure1 = Figure(figsize=(6,4), dpi=72)  
			self.axis1 = self.figure1.add_subplot(111)
		if self.graph1counter>0:
			self.axis1 = self.figure1.add_subplot(111) 
			self.axis1.clear()
			graphview.remove(self.canvas1)
			self.builder.get_object("toolbar1").remove(self.toolbar1)	
			
		self.befile=befile
		self.affile=affile
		if befile.get_filename()!=None:
			self.plot1(befile, "ro", "b-", "Before Illumination", widget)
		if affile.get_filename()!=None:
			self.plot1(affile, "go", "r-", "After Illumination", widget)
			
		self.axis1.set_xlabel('Minority Carrier Density (cm$^{-3}$)')   
		self.axis1.set_ylabel('Measured Lifetime (s)')   
		self.axis1.ticklabel_format(style='sci',scilimits=(0,0))
		self.axis1.grid(True)
		self.canvas1=FigureCanvasGTKAgg(self.figure1)
		self.canvas1.show()
		graphview.pack_start(self.canvas1, True, True)
		self.toolbar1 = NavigationToolbar(self.canvas1, self.builder.get_object("MainWindow"))
		self.builder.get_object("toolbar1").pack_start(self.toolbar1, False, False)
		self.scaletxtedit(widget)
		self.graph1counter+=1
		if befile.get_filename()!=None and affile.get_filename()!=None:
			self.checkresistivity(widget)
			self.builder.get_object("plbeforefile").set_sensitive(True)
			self.builder.get_object("recalccopbtn").set_sensitive(True)
			self.builder.get_object("calcdopefromres").set_sensitive(True)
			self.builder.get_object("restoredopingbtn").set_sensitive(True)
			self.builder.get_object("editfitbtn").set_sensitive(True)

	def plot1(self, file1, pointcol, linecol, mylabel, widget):
		#note there is no adjustment for extreme values here, although an adjustment of the limit method could perhaps provide this
		filename=file1.get_filename()
		fitpoints=[]
		datalist=excel.getValues(filename)
		tauvalues=datalist[1]
		deltan=datalist[2]
		if datalist==[0,0,0]:
			self.errorshow(self)
			return 1

		if deltan[0]>self.fitplotlimit:
			self.fitplotlimit=deltan[0]
		#print self.fitplotlimit
		#print isinstance( self.fitplotlimit, ( int, long, float ) )

		#limit = excel.getlocalmin(datalist[1])
		self.axis1.plot(deltan,tauvalues, pointcol, label=mylabel)

		#self.axis1.plot((deltan)[0:limit+1],(tauvalues)[0:limit+1], linecol, label=mylabel)
		
		#fitting
		#indices=datalist[2]>1E14
		#print (datalist[2])[indices]
		#print tauvalues[0:limit+1]
		#print deltan[0:limit+1]
		#fix this later

		range1=np.arange(0,self.fitplotlimit,0.5e14)
		#print (deltan[int(0.65*len(deltan))])
		offset=0.60
		#try:
		if mylabel=="Before Illumination":
			if self.beforeplotcounter==0:
				self.bfitrmin=deltan[int(offset*len(deltan))]
				self.bfitrmax=deltan[0]
				self.builder.get_object("bfitrmin").set_text("%.4g" % self.bfitrmin)
				self.builder.get_object("bfitrmax").set_text("%.4g" % self.bfitrmax)

			self.bcanfit=1
			#need to get from deltan values to actual offset
			#Find min:
			i=0
			self.beforeplotcounter+=1
			while deltan[i]>self.bfitrmax:
				i+=1
			actualmax=i
			i=len(deltan)-1
			while deltan[i]<self.bfitrmin:
				i-=1
			actualmin=i
			self.beforefitparams=fitting.fitting(np.array(deltan[actualmax:actualmin+1], np.float64), np.array(tauvalues[actualmax:actualmin+1], np.float64), self.startguessb)[0]
			#print fitting.fitting(np.array(deltan[actualmax:actualmin+1], np.float64), np.array(tauvalues[actualmax:actualmin+1], np.float64), self.startguessb)[1]
			i=actualmax
			squarediff=0
			taup, taun, n1, p1, NA=self.beforefitparams
			while i<actualmin+1:
				squarediff+=pow(tauvalues[i]-(((taup*n1)+(taup*deltan[i])+(taun*p1)+(taun*NA)+(taun*deltan[i]))/(NA+deltan[i])),2)
				i+=1
			self.builder.get_object("squarediffbefore").set_text("%.4g" % np.sqrt(squarediff))
			self.bsquarediff= np.sqrt(squarediff)
			if self.fitusebox.get_active()==0:
				self.axis1.plot(range1, (((taup*n1)+(taup*range1)+(taun*p1)+(taun*NA)+(taun*range1))/(NA+range1)), linecol, label="Before Illumination Fit")
				self.beforefitplot=[range1, (((taup*n1)+(taup*range1)+(taun*p1)+(taun*NA)+(taun*range1))/(NA+range1))]
				# print "Before:"
				# print "Trap from valence: %.4g" % ((1.3806503E-23*300*np.log(1.83E19/p1))/1.602E-19)
				# print "Trap from conduction: %.4g" % (((1.14*1.602E-19)-(1.3806503E-23*300*np.log(2.82E19/n1)))/1.6E-19)
				# print "Ec-Ev: %.4g" % ((1.3806503E-23*300*np.log((1.83E19*2.82E19)/(n1*p1)))/1.602E-19)
			elif self.fitusebox.get_active()==1:
				self.axis1.plot(deltan[0:int(offset*len(deltan))], tauvalues[0:int(offset*len(deltan))], linecol, label="Before Illumination Interpolation")
		elif mylabel=="After Illumination":
			if self.afterplotcounter==0:
				self.afitrmin=deltan[int(offset*len(deltan))]
				self.afitrmax=deltan[0]
				self.builder.get_object("afitrmin").set_text("%.4g" % self.afitrmin)
				self.builder.get_object("afitrmax").set_text("%.4g" % self.afitrmax)
			self.acanfit=1
			i=0
			self.afterplotcounter+=1
			while deltan[i]>self.afitrmax:
				#print "max: %.4g deltan: %.4g" % (self.afitrmax, deltan[i])
				i+=1
			actualmax=i
			i=len(deltan)-1
			while deltan[i]<self.afitrmin:
				i-=1
			actualmin=i
			self.afterfitparams=fitting.fitting(np.array(deltan[actualmax:actualmin], np.float64), np.array(tauvalues[actualmax:actualmin], np.float64), self.startguessa)[0]
			taup, taun, n1, p1, NA=self.afterfitparams
			i=actualmax
			squarediff=0
			while i<actualmin+1:
				squarediff+=pow(tauvalues[i]-(((taup*n1)+(taup*deltan[i])+(taun*p1)+(taun*NA)+(taun*deltan[i]))/(NA+deltan[i])),2)
				i+=1
			self.builder.get_object("squarediffafter").set_text("%.4g" % np.sqrt(squarediff))
			self.asquarediff= np.sqrt(squarediff)
			# print "After:"
			# print "Trap from valence: %.4g" % ((1.3806503E-23*300*np.log(1.83E19/p1))/1.602E-19)
			# print "Trap from conduction: %.4g" % (((1.14*1.602E-19)-(1.3806503E-23*300*np.log(2.82E19/n1)))/1.6E-19)
			# print "Ec-Ev: %.4g" % ((1.3806503E-23*300*np.log((1.83E19*2.82E19)/(n1*p1)))/1.602E-19)


			if self.fitusebox.get_active()==0:
				self.axis1.plot(range1, (((taup*n1)+(taup*range1)+(taun*p1)+(taun*NA)+(taun*range1))/(NA+range1)), linecol, label="After Illumination Fit")
				self.afterfitplot=[range1, (((taup*n1)+(taup*range1)+(taun*p1)+(taun*NA)+(taun*range1))/(NA+range1))]
				self.ironplottypebox.set_active(0)
				self.builder.get_object("maptypelabel").set_label("<b>Individual C values</b>")
			elif self.fitusebox.get_active()==1:
				self.axis1.plot(deltan[0:int(offset*len(deltan))], tauvalues[0:int(offset*len(deltan))], linecol, label="After Illumination Interpolation")


		self.axis1.legend(loc=4, prop={'size':10})
		# except:
		# 	if mylabel=="Before Illumination":
		# 		self.bcanfit=0
		# 	elif mylabel=="After Illumination":
		# 		self.acanfit=0
		# 	self.builder.get_object("fiterror").show()
		# 	self.canfit=0
		# 	self.ironplottypebox.set_active(1)
		# 	self.builder.get_object("maptypelabel").set_label("<b>C value from Generation Level</b>")
	def checkresistivity(self,widget):
		#remember the order of the constants
		#vthermal sigmani p1i sigmapi sigmanb n1b sigmapb
		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]
		befile=self.builder.get_object("filebeforebtn")
		affile=self.builder.get_object("fileafterbtn")
		belist=excel.getValues(befile.get_filename())
		aflist=excel.getValues(affile.get_filename())
		self.oldres=belist[0]
		self.dope=doping.calcDoping(self.oldres)
		self.olddope=self.dope
		self.builder.get_object("crossovertheory").set_text("%.4g" % concentration.COPcalc(self.constants, self.dope))
		self.builder.get_object("doping").set_text("%.4g" % self.dope)
		self.builder.get_object("pldopingtxt").set_text("%.4g" % self.dope)		
		if belist[0]!=aflist[0]:
			self.resbeforeval=belist[0]
			self.resafterval=aflist[0]
			self.builder.get_object("resbeforevallabel").set_label("%.4g" % self.resbeforeval)
			self.builder.get_object("resaftervallabel").set_label("%.4g" % self.resafterval)
			self.builder.get_object("reswindow").show()

		self.builder.get_object("resistivity").set_text("%.4g" % belist[0])


	def resbeforevalbtnclicked(self, widget):
		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]
		self.oldres=self.resbeforeval
		self.dope=doping.calcDoping(self.resbeforeval)
		self.olddope=self.dope
		self.builder.get_object("resistivity").set_text("%.4g" % self.resbeforeval)
		self.builder.get_object("pldopingtxt").set_text("%.4g" % self.dope)
		self.builder.get_object("doping").set_text("%.4g" % self.dope)
		self.builder.get_object("crossovertheory").set_text("%.4g" % concentration.COPcalc(self.constants, self.dope))		
		self.builder.get_object("reswindow").hide()

	def resaftervalbtnclicked(self, widget):
		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]
		self.dope=doping.calcDoping(self.resafterval)
		self.oldres=self.resafterval
		self.olddope=self.dope
		self.builder.get_object("resistivity").set_text("%.4g" % self.resafterval)
		self.builder.get_object("pldopingtxt").set_text("%.4g" % self.dope)
		self.builder.get_object("doping").set_text("%.4g" % self.dope)
		self.builder.get_object("crossovertheory").set_text("%.4g" % concentration.COPcalc(self.constants, self.dope))		
		self.builder.get_object("reswindow").hide()

	def resmyvalbtnclicked(self, widget):
		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]
		myres=float(self.builder.get_object("myrestxt").get_text())
		self.oldres=myres
		self.dope=doping.calcDoping(myres)
		self.olddope=self.dope
		self.builder.get_object("resistivity").set_text("%.4g" % myres)
		self.builder.get_object("pldopingtxt").set_text("%.4g" % self.dope)
		self.builder.get_object("doping").set_text("%.4g" % self.dope)	
		self.builder.get_object("crossovertheory").set_text("%.4g" % concentration.COPcalc(self.constants, self.dope))	
		self.builder.get_object("reswindow").hide()


	def recalccopbtnclicked(self, widget):
		#new doping concentration
		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]
		self.dope=float(self.builder.get_object("doping").get_text())
		self.builder.get_object("crossovertheory").set_text("%.4g" % concentration.COPcalc(self.constants, self.dope))
		self.builder.get_object("pldopingtxt").set_text("%.4g" % self.dope)
		#get new resistivity
		self.builder.get_object("resistivity").set_text("%.4g" % doping.calcRes(self.dope))
		

	def restoredopingbtnclicked(self, widget):
		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]
		self.dope=self.olddope
		self.builder.get_object("crossovertheory").set_text("%.4g" % concentration.COPcalc(self.constants, self.dope))
		self.builder.get_object("pldopingtxt").set_text("%.4g" % self.dope)
		self.builder.get_object("doping").set_text("%.4g" % self.dope)
		self.builder.get_object("resistivity").set_text("%.4g" % self.oldres)

	def errorclose(self, widget, response):
		widget.hide()
		#response is -7 for close
	
	def errorshow(self, widget):
		errorwin=self.builder.get_object("ErrorWindow")
		errorwin.show()

	def numbify(self, entry):
		text = entry.get_text().strip()

		entry.set_text(''.join([i for i in text if i in '.-eE0123456789']))

	        if gtk.Buildable.get_name(entry)=="doping":
		 	self.builder.get_object("pldopingtxt").set_text("%.4g" % float(entry.get_text()))

	def calcdopefromresistivity(self, widget):
		self.dope=doping.calcDoping(float(self.builder.get_object("resistivity").get_text()))
		self.builder.get_object("crossovertheory").set_text("%.4g" % concentration.COPcalc(self.constants, self.dope))
		self.builder.get_object("pldopingtxt").set_text("%.4g" % self.dope)
		self.builder.get_object("doping").set_text("%.4g" % self.dope)
		
	def scaletxtedit(self, widget):

		if widget.get_name()=="GtkEntry":
			self.numbify(widget)

		if (self.builder.get_object("filebeforebtn").get_filename()!=None or self.builder.get_object("fileafterbtn").get_filename()!=None):
			self.builder.get_object("scalebtn").set_sensitive(True)
		else:
			self.builder.get_object("scalebtn").set_sensitive(False)
		
		if self.builder.get_object("filebeforebtn").get_filename()!=None or self.builder.get_object("fileafterbtn").get_filename()!=None:
			self.builder.get_object("resetbtn").set_sensitive(True)
		else:
			self.builder.get_object("resetbtn").set_sensitive(False)

		if self.builder.get_object("filebeforebtn").get_filename()!=None and self.builder.get_object("fileafterbtn").get_filename()!=None:
			self.builder.get_object("ironcalcbtn").set_sensitive(True)
			self.builder.get_object("cfactorbtn").set_sensitive(True)
		else:
			self.builder.get_object("ironcalcbtn").set_sensitive(False)
			self.builder.get_object("cfactorbtn").set_sensitive(False)

		if self.builder.get_object("plbeforefile").get_filename()!=None and self.builder.get_object("plafterfile").get_filename()!=None  and self.builder.get_object("filebeforebtn").get_filename()!=None and self.builder.get_object("fileafterbtn").get_filename()!=None:
			self.builder.get_object("plcalcbtn").set_sensitive(True)
		else:
			self.builder.get_object("plcalcbtn").set_sensitive(False)

	def scalebtnclicked(self, widget):
		graphview=self.builder.get_object("graph1")

		#This sometimes causes some strange error asking for 2-tuples, it might be a bug in Cygwin

		xmin1=self.builder.get_object("xmin1").get_text()
		xmax1=self.builder.get_object("xmax1").get_text()
		ymin1=self.builder.get_object("ymin1").get_text()
		ymax1=self.builder.get_object("ymax1").get_text()
		xaxis=self.axis1.get_xlim()
		yaxis=self.axis1.get_ylim()
		xmaxmag="%e" % xaxis[1]		
		ymaxmag="%e" % yaxis[1]		
		xmaxmag=xmaxmag[xmaxmag.index("e")+1:len(xmaxmag)]
		ymaxmag=ymaxmag[ymaxmag.index("e")+1:len(ymaxmag)]

		if xmin1 != "" and xmin1 !=None:
			if xmin1.find("e")==-1 and xmin1.find("E")==-1:
				xmin1=float(xmin1)*pow(10, float(xmaxmag))
			self.axis1.set_xlim(left=float(xmin1))

		if xmax1 != "" and xmax1 != None:
			if xmax1.find("e")==-1 and xmax1.find("E")==-1:
				xmax1=float(xmax1)*pow(10, float(xmaxmag))
			self.axis1.set_xlim(right=float(xmax1))

		if ymax1 != "" and ymax1 != None:
			if ymax1.find("e")==-1 and ymax1.find("E")==-1:
				ymax1=float(ymax1)*pow(10, float(ymaxmag))
			self.axis1.set_ylim(top=float(ymax1))
			
		if ymin1 != "" and ymin1 != None:
			if ymin1.find("e")==-1 and ymin1.find("E")==-1:
				ymin1=float(ymin1)*pow(10, float(ymaxmag))
			self.axis1.set_ylim(bottom=float(ymin1))

		yaxis=self.axis1.get_ylim()
		xaxis=self.axis1.get_xlim()
		if yaxis[0]>yaxis[1] or xaxis[0]>xaxis[1]:
			self.builder.get_object("plotflippeddialog").show()

		graphview.remove(self.canvas1)	
		graphview.pack_start(self.canvas1, True, True)

	# def ironcheckbuttons(self, widget):


	# 	if widget.get_name()=="GtkEntry":
	# 		self.numbify(widget)

#I'm not sure if this function is still necessary will check what uses it later.

	def scalebtn2clicked(self, widget):
		graphview=self.builder.get_object("irongraph")
		xmin2=self.builder.get_object("xmin2").get_text()
		xmax2=self.builder.get_object("xmax2").get_text()
		ymin2=self.builder.get_object("ymin2").get_text()
		ymax2=self.builder.get_object("ymax2").get_text()
		xaxis=self.axis2.get_xlim()
		yaxis=self.axis2.get_ylim()
		xmaxmag="%e" % xaxis[1]		
		ymaxmag="%e" % yaxis[1]		
		xmaxmag=xmaxmag[xmaxmag.index("e")+1:len(xmaxmag)]
		ymaxmag=ymaxmag[ymaxmag.index("e")+1:len(ymaxmag)]
		
		if xmax2!="" and xmax2 != None:
			if xmax2.find("e")==-1 and xmax2.find("E")==-1:
				xmax2=float(xmax2)*pow(10, float(xmaxmag))
			self.axis2.set_xlim(right=float(xmax2))
		if xmin2!="" and xmin2 != None:
			if xmin2.find("e")==-1 and xmin2.find("E")==-1:
				xmin2=float(xmin2)*pow(10, float(xmaxmag))
			self.axis2.set_xlim(left=float(xmin2))
		if ymax2!="" and ymax2!=None:
			if ymax2.find("e")==-1 and ymax2.find("E")==-1:
				ymax2=float(ymax2)*pow(10, float(ymaxmag))
			self.axis2.set_ylim(top=float(ymax2))
		if ymin2!="" and ymin2!=None:
			if ymin2.find("e")==-1 and ymin2.find("E")==-1:
				ymin2=float(ymin2)*pow(10, float(ymaxmag))
			self.axis2.set_ylim(bottom=float(ymin2))

		xaxis=self.axis2.get_xlim()
		yaxis=self.axis2.get_ylim()

		if xaxis[0]>xaxis[1] or yaxis[0]>yaxis[1]:
			self.builder.get_object("plotflippeddialog").show()





		graphview.remove(self.canvas2)	
		graphview.pack_start(self.canvas2, True, True)				

	def ironcalcbtnclicked(self, widget):
		taupb, taunb, n1b, p1b, NAb=self.beforefitparams
		taupa, tauna, n1a, p1a, NAa=self.afterfitparams
		a=(taupb+taunb-taupa-tauna)
		b=((NAa*taupb)+(NAa*taunb)+(n1b*taupb)+(p1b*taunb)+(taunb*NAb)-(NAb*taupa)-(NAb*tauna)-(n1a*taupa)-(tauna*p1a)-(tauna*NAa))
		c=(NAa*n1b*taupb)+(NAa*taunb*p1b)+(NAa*taunb*NAb)-(NAb*n1a*taupa)-(NAb*tauna*p1a)-(NAb*tauna*NAa)
		copfit1= (-b+np.sqrt(pow(b,2)-(4*a*c)))/(2*a)
		copfit2= (-b-np.sqrt(pow(b,2)-(4*a*c)))/(2*a)
		self.builder.get_object("copfit1").set_label("%.4g" % copfit1)
		self.builder.get_object("copfit2").set_label("%.4g" % copfit2)

		self.plotgraph1(widget)
		self.dope=float(self.builder.get_object("doping").get_text())
		beforelist=excel.getValues(self.builder.get_object("filebeforebtn").get_filename())
		afterlist=excel.getValues(self.builder.get_object("fileafterbtn").get_filename())
		if len(afterlist[1])!=len(afterlist[2]) or len(beforelist[1])!=len(beforelist[2]) or len(afterlist[1])!=len(beforelist[1]):
			print "ERROR"
			#add proper error handling later
		beforelist[2].reverse()
		afterlist[2].reverse()
		beforelist[1].reverse()
		afterlist[1].reverse()
		
		self.bqsspclife=beforelist
		self.aqsspclife=afterlist
		if self.fitusebox.get_active()==0:
			bdict=dict(zip(beforelist[2], beforelist[1]))
			adict=dict(zip(afterlist[2], afterlist[1]))
			i=0
			tafterfitdict={}
			tbeforefitdict={}
			while i<len(beforelist[2]):
				taup, taun, n1, p1, NA=self.afterfitparams
				tafterfitdict[beforelist[2][i]]=(((taup*n1)+(taup*beforelist[2][i])+(taun*p1)+(taun*NA)+(taun*beforelist[2][i]))/(NA+beforelist[2][i]))
				taup, taun, n1, p1, NA=self.beforefitparams
				tbeforefitdict[afterlist[2][i]]=(((taup*n1)+(taup*afterlist[2][i])+(taun*p1)+(taun*NA)+(taun*afterlist[2][i]))/(NA+afterlist[2][i]))
				i+=1
			totalbeforeplotdict=dict(tbeforefitdict.items()+bdict.items())
			totalafterplotdict=dict(tafterfitdict.items()+adict.items())
			dictlist=[totalbeforeplotdict, totalafterplotdict]
				
		elif self.fitusebox.get_active()==1:
			dictlist=concentration.interpolation(beforelist, afterlist)
		#dictlist[0] is before, dictlist[1] is after values

		for key in dictlist[0].keys():
			if dictlist[1].has_key(key)!=True:
				del dictlist[0][key]

		for key in dictlist[1].keys():
			if dictlist[0].has_key(key)!=True:
				del dictlist[1][key]

		if len(dictlist[0])!=len(dictlist[1]):
			print "dict lengths not equal"
			#add proper error handling later
			#difference is expected but must be handled
		
		ironvalues=[]
		qsspccvals=[]


		sortedbkeylist=sorted(dictlist[0].iterkeys())
		sortedakeylist=sorted(dictlist[1].iterkeys())
		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]

		for bkey in sortedbkeylist:
			C=concentration.calcPrefactor(self.constants, self.dope, bkey)
			iron=1E-6*(concentration.calcFeConc(C, dictlist[0][bkey], dictlist[1][bkey]))
			ironvalues.append(iron)
			qsspccvals.append(C)
			#print "iron: %.4g, tn0: %.4g, tp0: %.4g, doping: %.4g, tbefore: %.4g, tafter: %.4g, deltan: %.4g \n" % (iron, tn0, tp0, self.dope, dictlist[0][bkey], dictlist[1][bkey], bkey)


		xlim=self.axis1.get_xlim()
		ylim=self.axis1.get_ylim()

		self.axis1.set_xlim(xlim)
		self.axis1.set_ylim(ylim)

		graphview = self.builder.get_object("irongraph")  
		if self.ironcounter==0:
			self.figure2 = Figure(figsize=(6,4), dpi=72)  
			self.axis2 = self.figure2.add_subplot(111)
		if self.ironcounter>0:
			self.axis2 = self.figure2.add_subplot(111) 
			self.axis2.clear()
			graphview.remove(self.canvas2)
			self.builder.get_object("toolbar2").remove(self.toolbar2)

		#self.axis2.plot(sortedbkeylist, ironvalues, "b-") #maybe make this within limit later		
		if self.fitusebox.get_active()==0:
			i=0
			ironvals=[]
			self.fitcvals=[]
			while i<len(self.beforefitplot[0]):
				C=concentration.calcPrefactor(self.constants, self.dope, self.beforefitplot[0][i])
				iron=1E-6*(concentration.calcFeConc(C, self.beforefitplot[1][i], self.afterfitplot[1][i]))
				self.fitcvals.append(C)
				ironvals.append(iron)
				i+=1
			self.axis2.plot(self.beforefitplot[0], ironvals, "b-")
			self.associron=ironvals
		else:
			self.axis2.plot(sortedbkeylist,ironvalues, "ro")
					   
					   
		self.irondata=[sortedbkeylist, ironvalues, qsspccvals]
		self.axis2.set_xlabel('Minority Carrier Density (cm$^{-3}$)')   
		self.axis2.set_ylabel('Interstitial Iron Concentration (cm$^{-3}$)') #check units are correct   
		self.axis2.ticklabel_format(style='sci',scilimits=(0,0))
		self.axis2.grid(True)
		self.canvas2=FigureCanvasGTKAgg(self.figure2)
		self.canvas2.show()
		graphview.pack_start(self.canvas2, True, True)
		self.toolbar2 = NavigationToolbar(self.canvas2, self.builder.get_object("MainWindow"))
		self.builder.get_object("toolbar2").pack_start(self.toolbar2, False, False)
		self.ironcounter+=1
		self.builder.get_object("resetbtn2").set_sensitive(True)

		cop=(sortedbkeylist[ironvalues.index(max(ironvalues))] + sortedbkeylist[ironvalues.index(min(ironvalues))])/2
		self.builder.get_object("crossovertxt").set_text("%.4g" % cop)
		self.builder.get_object("savedatabtn").set_sensitive(True)
		self.builder.get_object("getfebtn").set_sensitive(True)
	


		#don't delete just change default view
		ironmean=np.mean(ironvalues[-10:-1])
		ironmin=max(ironvalues)
		i=ironvalues.index(ironmin)
		stddevfe=np.std(ironvalues)
		while i<len(ironvalues):
			#0.3 seems good compromise factor
			if abs(ironvalues[i])< abs(ironmean+(0.3*np.std(ironvalues))):
				ironmin=ironvalues[i]
				break
			i+=1
			
		#set default view limits
		self.axis2.set_ylim(bottom=1E10)
		self.axis2.set_ylim(top=ironmin)

		self.builder.get_object("feilabel").set_label("[Fe<sub>i</sub>] (cm<sup>-3</sup>): ")
		self.builder.get_object("getfebtnlabel").set_label("Get [Fe<sub>i</sub>]")
		self.curplot="Fe"
		self.builder.get_object("deltangettxt").set_text("")
		self.builder.get_object("fegettxt").set_text("")
		self.builder.get_object("scalebtn2").set_sensitive(True)
		self.qsspclist=[sortedbkeylist, ironvalues]
		self.builder.get_object("meancalcbtn").set_sensitive(True)
		self.meanmode="Fe"
		if self.builder.get_object("meanmin").get_text()==None or self.builder.get_object("meanmin").get_text()=="" or self.builder.get_object("meanmax").get_text()==None or self.builder.get_object("meanmin").get_text()=="":
			self.builder.get_object("meanmin").set_text("%.4g" % sortedbkeylist[-10])
			self.builder.get_object("meanmax").set_text("%.4g" % sortedbkeylist[-1])

		# ironmean=np.mean(ironvalues[-10:-1])
		# self.qsspcironmeanlast=ironmean
		# self.builder.get_object("meanconctxt").set_text("%.4g" % ironmean)
		self.builder.get_object("meanlabel").set_label("Mean of the interstitial\nIron Concentration in\nthe range given (cm<sup>-3</sup>): ")
		self.builder.get_object("meancalcbtn").set_label("Calculate mean Iron Concentration\n              in the range given")
		self.calculatemeanrangeclicked(widget)

	def calculatemeanrangeclicked(self, widget):
		try:
			rangemin=float(self.builder.get_object("meanmin").get_text())
			rangemax=float(self.builder.get_object("meanmax").get_text())
			if self.meanmode=="Fe":
				i=0
				while self.qsspclist[0][i]<rangemin and i<len(self.qsspclist[0]):
					i+=1
				actualmin=i
				i=len(self.qsspclist[0])-1
				while self.qsspclist[0][i]>rangemax and i>=0:
					i-=1
				actualmax=i
				ironmean=np.mean(self.qsspclist[1][actualmin:actualmax+1])
				self.qsspcstdev=np.std(self.qsspclist[1][actualmin:actualmax+1])
				self.qsspcironmeanlast=ironmean
				self.builder.get_object("meanconctxt").set_text("%.4g" % ironmean)

			elif self.meanmode=="C":
				i=0
				while self.cvaluedn[i]<rangemin and i<len(self.cvaluedn):
					i+=1
				actualmin=i
				i=len(self.cvaluedn)-1
				while self.cvaluedn[i]>rangemax and i>=0:
					i-=1
				actualmax=i
				cmean=np.mean(self.cvalues[actualmin:actualmax+1])
				self.cmean=cmean
				self.cstdev=np.std(self.cvalues[actualmin:actualmax+1])
				self.builder.get_object("meanconctxt").set_text("%.4g" % cmean)
		except (ValueError, ZeroDivisionError, IndexError):
			self.builder.get_object("meanrangeerror").show()
			

	def plotdialogcancelbtnclicked(self,widget):
		self.builder.get_object("plotfilesavedialog").hide()
	def plotdialogsavebtnclicked(self, widget):
		self.cursave="plotdialogsave"
		filename=self.builder.get_object("plotfilesavedialog").get_filename()
		if filename[-4:len(filename)]!=".png" and filename[-4:len(filename)]!=".PNG":
			filename += ".png"
		self.currentfilename=filename

		if os.path.exists(filename)==True:
			self.builder.get_object("overwritewarning").show()
		else:
			self.figure2.savefig(filename, format="png")
		#add error handling
		#check if filename has extension, if not append PNG - maybe ask?

		# note that this inherently uses current scaling options - WYSIWYG
		
	def savedatabtnclicked(self,widget):
		self.builder.get_object("datafilesavedialog").show()
		self.datasave="plot"

	def datadialogcancelbtnclicked(self,widget):
		self.builder.get_object("datafilesavedialog").hide()	
	
	def datadialogsavebtnclicked(self,widget):
		filename=self.builder.get_object("datafilesavedialog").get_filename()

		if filename[-4:len(filename)]!=".txt" and filename[-4:len(filename)]!=".TXT":
			filename += ".txt"
		self.currentfilename=filename 
		if self.datasave=="plot":
			self.cursave="feplotdata"
		if self.datasave=="adjusteddata":
			self.cursave="adjusteddata"
		elif self.datasave=="map":
			self.cursave="femapdata"
		if os.path.exists(filename)==True:
			self.builder.get_object("overwritewarning").show()
		else:
			if self.datasave=="plot":
				self.datatxtsave()
			elif self.datasave=="map":
				if self.ironplottypebox.get_active()==0:
				#Individual C values from gen level, tau values
					savetxt(self.currentfilename, self.ironconcmatrixindgen, fmt="%12.6G")				
				elif self.ironplottypebox.get_active()==1:
				#Individual C values from fits
					savetxt(self.currentfilename, self.ironconcmatrixindividual, fmt="%12.6G")
				elif self.ironplottypebox.get_active()==2:
				#C value from Generation Level
					savetxt(self.currentfilename, self.ironconcmatrixgenlevel, fmt="%12.6G")
			elif self.ironplottypebox.get_active()==3:
					savetxt(self.currentfilename, self.ironconcmatrixmeanC, fmt="%12.6G")
				#Use mean of last C values

			elif self.datasave=="adjusteddata":
				savetxt(self.currentfilename, self.tauafter, fmt="%12.6G")

		self.builder.get_object("datafilesavedialog").hide()	
	
	
		#add error handling

	def datatxtsave(self):
		#QSSPC data self.bqsspclife, self.aqsspclife tau, n 1,2 and resistivity is 0
		#Fitting formula and parameters and least squares
		#Fitted data: deltan, taubefore, tauafter
		#Iron concentration data
		self.builder.get_object("datafilesavedialog").hide()
		textfile=open(self.currentfilename, "w")
		textfile.write("Constants: vthermal: %.4g\tsigmani: %.4g\tp1i: %.4g\tsigmapi: %.4g\tsigmanb: %.4g\tn1b: %.4g\tsigmapb: %.4g\n" % (self.constants[0],self.constants[1],self.constants[2],self.constants[3],self.constants[4],self.constants[5],self.constants[6]))
		textfile.write("QSSPC Data:\n")
		textfile.write("Before Illumination: Resistivity=%.4g:\n" % self.bqsspclife[0])
		textfile.write("deltaN(cm^-3)\tTau before(s)\n")
		i=0
		while i<len(self.bqsspclife[2]):
			textfile.write("%.4g\t%.4g\n" % (self.bqsspclife[2][i], self.bqsspclife[1][i]))
			i+=1
		textfile.write("\nAfter Illumination: Resistivity=%.4g:\n" % self.aqsspclife[0])
		i=0
		while i<len(self.aqsspclife[2]):
			textfile.write("%.4g\t%.4g\n" % (self.aqsspclife[2][i], self.aqsspclife[1][i]))
			i+=1

		textfile.write("\nFitting Data:\n")
		textfile.write("Fitting equation: (((taup*n1)+(taup*deltaN)+(taun*p1)+(taun*NA)+(taun*deltaN))/(NA+deltaN))\n")
		taup, taun, n1, p1, NA=self.beforefitparams
		textfile.write("Before Illumination Fit data:\n")
		textfile.write("Before Illumination Fitting range (in deltaN): Minimum: %.4g\tMaximum: %.4g\n" % (self.bfitrmin, self.bfitrmax) )
		textfile.write("Before Illumination Fit parameters: taup: %.4g\ttaun: %.4g\tn1: %.4g\tp1: %.4g\tNA: %.4g\n" % (taup, taun, n1, p1, NA))
		textfile.write("Square root of the sum of the squared differences within the fitting range for the Before Illumination fit: %.4g\n" % self.bsquarediff)
		textfile.write("After Illumination Fit data:\n")
		textfile.write("After Illumination Fitting range (in deltaN): Minimum: %.4g\tMaximum: %.4g\n" % (self.afitrmin, self.afitrmax) )
		taup, taun, n1, p1, NA=self.afterfitparams
		textfile.write("After Illumination Fit parameters: taup: %.4g\ttaun: %.4g\tn1: %.4g\tp1: %.4g\tNA: %.4g\n" % (taup, taun, n1, p1, NA))
		textfile.write("Square root of the sum of the squared differences within the fitting range for the Before Illumination fit: %.4g\n" % self.asquarediff)
		textfile.write("Plotted fit values for Before and After Illumination:\n")
		textfile.write("deltaN(cm^-3)\tTau before (s)\tTau after(s)\tPrefactor value\tInterstitial Iron Concentration (cm^-3)\n")
		i=0
		while i<len(self.beforefitplot[0]):
			textfile.write("%.4g\t%.4g\t%.4g\t%.4g\t%.4g\n" % (self.beforefitplot[0][i], self.beforefitplot[1][i], self.afterfitplot[1][i], self.fitcvals[i], self.associron[i]))
			i+=1
		
		textfile.write("\nQSSPC Iron Concentration data:\n")
		textfile.write("deltaN(cm^-3)\tPrefactor\t[Fe](cm^-3)\n")
		i=0

		while i<len(self.irondata[0]):
			textfile.write("%.4g\t%.4g\t%.4g\n" % (self.irondata[0][i], self.irondata[2][i], self.irondata[1][i]))
			i+=1
		textfile.close()

		if self.delimbox.get_active()==1:
			textfile=open(self.currentfilename, "r")
			buffer=textfile.read()
			textfile.close()
			buffer=buffer.replace(".",",")
			textfile=open(self.currentfilename, "w")
			textfile.write(buffer)
			textfile.close()


		
	def cfactorbtnclicked(self, widget):
		self.plotgraph1(widget)
		self.dope=float(self.builder.get_object("doping").get_text())
		beforelist=excel.getValues(self.builder.get_object("filebeforebtn").get_filename())
		afterlist=excel.getValues(self.builder.get_object("fileafterbtn").get_filename())
		if len(afterlist[1])!=len(afterlist[2]) or len(beforelist[1])!=len(beforelist[2]) or len(afterlist[1])!=len(beforelist[1]):
			print "ERROR"
			#add proper error handling later
		beforelist[2].reverse()
		afterlist[2].reverse()
		beforelist[1].reverse()
		afterlist[1].reverse()
		dictlist=concentration.interpolation(beforelist, afterlist)
		#dictlist[0] is before, dictlist[1] is after values

		for key in dictlist[0].keys():
			if dictlist[1].has_key(key)!=True:
				del dictlist[0][key]

		for key in dictlist[1].keys():
			if dictlist[0].has_key(key)!=True:
				del dictlist[1][key]

		if len(dictlist[0])!=len(dictlist[1]):
			print "dict lengths not equal"
			#add proper error handling later
			#difference is expected but must be handled
		
		cvalues=[]

		sortedbkeylist=sorted(dictlist[0].iterkeys())
		sortedakeylist=sorted(dictlist[1].iterkeys())
		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]
		for bkey in sortedbkeylist:
			cvalues.append(concentration.calcPrefactor(self.constants, self.dope, bkey))

		#plot cvalues
		graphview = self.builder.get_object("irongraph")  
		if self.ironcounter==0:
			self.figure2 = Figure(figsize=(6,4), dpi=72)  
			self.axis2 = self.figure2.add_subplot(111)
		if self.ironcounter>0:
			self.axis2 = self.figure2.add_subplot(111) 
			self.axis2.clear()
			graphview.remove(self.canvas2)
			self.builder.get_object("toolbar2").remove(self.toolbar2)
		self.axis2.plot(sortedbkeylist,cvalues, "bo")
		#self.axis2.plot(sortedbkeylist, ironvalues, "b-") #maybe make this within limit later		 

		if self.fitusebox.get_active()==0:
			i=0
			ironvals=[]
			self.fitcvals=[]
			while i<len(self.beforefitplot[0]):
				C=concentration.calcPrefactor(self.constants, self.dope, self.beforefitplot[0][i])
				self.fitcvals.append(C)
				i+=1
			self.axis2.plot(self.beforefitplot[0], self.fitcvals, "r-")

		self.axis2.set_xlabel('Minority Carrier Density (cm$^{-3}$)')   
		self.axis2.set_ylabel('Prefactor, C') #check units are correct   
		self.axis2.ticklabel_format(style='sci',scilimits=(0,0))
		self.axis2.grid(True)
		self.canvas2=FigureCanvasGTKAgg(self.figure2)
		self.canvas2.show()
		graphview.pack_start(self.canvas2, True, True)
		self.toolbar2 = NavigationToolbar(self.canvas2, self.builder.get_object("MainWindow"))
		self.builder.get_object("toolbar2").pack_start(self.toolbar2, False, False)
		self.ironcounter+=1
		self.builder.get_object("resetbtn2").set_sensitive(True)
		self.builder.get_object("savedatabtn").set_sensitive(False)

		self.builder.get_object("feilabel").set_label("C: ")
		self.builder.get_object("getfebtnlabel").set_label("Get C")
		

		self.cmean= np.mean(cvalues[-10:-1])
		# self.builder.get_object("meanconctxt").set_text("%.4g" % self.cmean)
		self.curplot="C"

		self.cvalues=cvalues
		self.cvaluedn=sortedbkeylist
		#set default view
		#find ymin
		i=cvalues.index(max(cvalues))+1
		
		while i<len(cvalues):
			if cvalues[i]> (self.cmean-(abs(self.cmean)*4)):
				self.axis2.set_ylim(bottom=cvalues[i])
				break
			else:
				i+=1

		self.axis2.set_ylim(top=self.cmean+(abs(self.cmean)*4))
		self.builder.get_object("deltangettxt").set_text("")
		self.builder.get_object("fegettxt").set_text("")
		self.builder.get_object("scalebtn2").set_sensitive(True)
		self.builder.get_object("meancalcbtn").set_sensitive(True)
		self.meanmode="C"
		self.builder.get_object("meanlabel").set_label("Mean of the Prefactor value\n   in the given range: ")	
		if self.builder.get_object("meanmin").get_text()==None or self.builder.get_object("meanmin").get_text()=="" or self.builder.get_object("meanmax").get_text()==None or self.builder.get_object("meanmin").get_text()=="":
			self.builder.get_object("meanmin").set_text("%.4g" % self.cvaluedn[-10])
			self.builder.get_object("meanmax").set_text("%.4g" % self.cvaluedn[-1])
		self.builder.get_object("meancalcbtn").set_label("Calculate mean Prefactor value\n      in the range given")
		self.calculatemeanrangeclicked(widget)

	
	def resetbtn2clicked(self, widget):
		if self.curplot=="C":
			self.cfactorbtnclicked(widget)
		elif self.curplot=="Fe":
			self.ironcalcbtnclicked(widget)

			
	def restoredefaultsbtnclicked(self, widget):
		self.builder.get_object("vthermaltxt").set_text("1.1E7")
		self.builder.get_object("sigmanitxt").set_text("1.3E-14")
		self.builder.get_object("p1itxt").set_text("1.28E13")
		self.builder.get_object("sigmapitxt").set_text("7E-17")
		self.builder.get_object("sigmanbtxt").set_text("5E-15")
		self.builder.get_object("n1btxt").set_text("1.22571E15")
		self.builder.get_object("sigmapbtxt").set_text("3E-15")
		
		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]

	def plfileset(self, widget):
		#self.
		if widget.get_name()=="GtkEntry":
			self.numbify(widget)

		if gtk.Buildable.get_name(widget)=="plafterfile":
			self.adjusted=0

		if self.builder.get_object("plbeforefile").get_filename()!=None and self.builder.get_object("plafterfile").get_filename()!=None and self.builder.get_object("filebeforebtn").get_filename()!=None and self.builder.get_object("fileafterbtn").get_filename()!=None and self.builder.get_object("pldopingtxt").get_text()!=None:
			self.builder.get_object("plcalcbtn").set_sensitive(True)
		else:
			self.builder.get_object("plcalcbtn").set_sensitive(False)

		if gtk.Buildable.get_name(widget)=="plbeforefile":
			#display dialog asking for generation level
			self.builder.get_object("pl2genleveldialog").show()
		
	def plcalcbtnclicked(self, widget):
		#try:

		if self.bcanfit==1 and self.acanfit==1:
			self.canfit=1
		else:
			self.canfit=0

		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]

		plbeforefile = open(self.builder.get_object("plbeforefile").get_filename(), "r")
		beforeraw = loadtxt(plbeforefile, dtype=float)
		self.taubefore = 1E6*beforeraw
		b = 1E6*beforeraw
		plbeforefile.close()
	       	#savetxt("beforefilesave.txt", self.taubefore, fmt="%12.6G")
		if self.adjusted==0:
			plafterfile = open(self.builder.get_object("plafterfile").get_filename(), "r")
			afterraw=loadtxt(plafterfile, dtype=float)
			self.tauafter=1E6*afterraw
			a=1E6*afterraw
			plafterfile.close()
		else:
			a=np.copy(self.tauafter)
		#accounts for some files using already microseconds
		if self.taubefore.mean()>1000:
			self.taubefore=self.taubefore/1E6
			b=b/1E6
		if self.tauafter.mean()>1000:
			self.tauafter=self.tauafter/1E6
			a=a/1E6
		# 1E6 factor is necessary to put in microseconds
		if np.min(self.taubefore)<0:
			indicesb=[]
			zeroarray=[]
			self.negarrayb=[]			
			oldshape=self.taubefore.shape
			flatbefore=self.taubefore.flatten()
			i=0
			while i<flatbefore.size:
				if flatbefore[i]<0:
					indicesb.append(i)
					zeroarray.append(0)
					self.negarrayb.append(-5E30)					
				i+=1

			flatbefore.put(indicesb, zeroarray)
			self.taubefore=flatbefore.reshape(oldshape)

		if np.min(self.tauafter)<0:
			indicesa=[]
			zeroarray=[]
			self.negarraya=[]
			oldshape=self.tauafter.shape
			flatafter=self.tauafter.flatten()
			i=0
			while i<flatafter.size:
				if flatafter[i]<0:
					indicesa.append(i)
					zeroarray.append(0)
					self.negarraya.append(-5E30)
				i+=1

			flatafter.put(indicesa, zeroarray)
			self.tauafter=flatafter.reshape(oldshape)
			
		self.indicesb=indicesb
		self.indicesa=indicesa
		# if np.min(self.taubefore)<0:
		# 	i=0
		# 	j=0
		# 	while i<1000:
		# 		while j<1000:
		# 			if self.taubefore[i][j]<0:
		# 				indicesb.append([i, j])
		# 				#self.taubefore[i][j]=0
		# 			j+=1
		# 		j=0
		# 		i+=1
		# i=0
		# while i<len(indicesb):
		# 	zeroarray.append(0)
		# 	i+=1

		# np.put(self.taubefore, indicesb, zeroarray)
		# indicesa=[]
		# zeroarray2=[]
		# if np.min(self.tauafter)<0:
		# 	i=0
		# 	j=0
		# 	while i<1000:
		# 		while j<1000:
		# 			if self.taubefore[i][j]<0:
		# 				#self.taubefore[i][j]=0
		# 				indicesa.append([i,j])
		# 			j+=1
		# 		j=0
		# 		i+=1
		# i=0
		# while i<len(indicesb):
		# 	zeroarray.append(0)
		# 	i+=1	
		# np.put(self.tauafter, indicesa, zeroarray)
		

		#Need injection level for genlevel iron values
		injectionlevel=float(self.builder.get_object("injectionleveltxt").get_text())
		C=concentration.calcPrefactor(self.constants, float(self.builder.get_object("pldopingtxt").get_text()), injectionlevel)
		self.ironconcmatrixgenlevel = abs(C*(1/self.taubefore - 1/self.tauafter))

		
		#HACK TO MAKE NEGATIVE VALUES POSITIVE
		e=abs(C*(1/self.taubefore - 1/self.tauafter))
		#mean C value iron values:
		self.cfactorbtnclicked(widget)
		self.ironconcmatrixmeanC = abs(self.cmean*(1/self.taubefore - 1/self.tauafter))

		
		self.deltanbindgen=self.genlevelbefore*1E-6*self.taubefore
		self.deltanaindgen=self.genlevelafter*1E-6*self.tauafter
		#worry about warning level shit later
		self.diffmatrixindgen=np.abs(self.deltanaindgen-self.deltanbindgen)
		print np.mean(self.diffmatrixindgen[100:-100,100:-100])
		nmean=(self.deltanaindgen+self.deltanbindgen)/2.0
		
		C=concentration.calcPrefactor(self.constants, float(self.builder.get_object("pldopingtxt").get_text()), nmean)				
		self.ironconcmatrixindgen=np.abs(C*(1/self.taubefore - 1/self.tauafter))
		#do warn level shit
		oldshape=self.ironconcmatrixindgen.shape
		ironflat=self.ironconcmatrixindgen.flatten()
		diffmatrixflatten=self.diffmatrixindgen.flatten()
		indicesb=[]
		indices=[]
		zeroarray=[]
		zeroarray2=[]
		count=0
		i=0
		matrixmean= np.mean(self.diffmatrixindgen[100:-100,100:-100])
		matrixstd= np.std(self.diffmatrixindgen[100:-100,100:-100])		
		while i<diffmatrixflatten.size:
			if diffmatrixflatten[i]>self.ddnlimit:
				indicesb.append(i)
				zeroarray2.append(-7E30)
				count+=1				
			elif ironflat[i]==np.Inf or ironflat[i]==-np.Inf or np.isnan(ironflat[i]):
				indices.append(i)
				zeroarray.append(-6E30)					
			i+=1
		#mean
		ironflat.put(indices, zeroarray)
		ironflat.put(indicesb, zeroarray2)
		self.ironconcmatrixindgen=ironflat.reshape(oldshape)
		if self.ironplottypebox.get_active()==0:
			 percent=(float(count)/diffmatrixflatten.size)*100
			 self.builder.get_object("warnmeanlabel").set_label("The warning level is %.4g.\nThe number of values above this was %i out of %i pixels.\nThe percentage of values with differences above this was %.4g%%. \nThe mean of the array of differences of injection level values was %.4g.\nThe standard deviation of the array of differences of the injection level was %.4g\nThese values will be set to -7E30 on the PL map (and so coloured differently).\nThe warning level can be changed in the Edit Plot Options window.\nYou must recalculate the maps after changing this though." % (self.ddnlimit, count, diffmatrixflatten.size, percent, matrixmean, matrixstd))
			 self.builder.get_object("warnpixlabel").set_label("%i/%i (%.4g%%)" % (count, diffmatrixflatten.size, percent))
			 self.builder.get_object("validpointstxt").set_label("%i/%i (%.4g%%)" % (diffmatrixflatten.size-count, diffmatrixflatten.size, 100.0-percent))		
			 self.builder.get_object("meanarraydifflabel").set_label("%.4g" % matrixmean)
			 self.builder.get_object("stdarraydifflabel").set_label("%.4g" % matrixstd)
			 self.builder.get_object("qsspcmeanwarning").show()
			 
		if self.canfit==1:
			self.ironconcmatrixindividual = np.ndarray(self.tauafter.shape, dtype=np.float)
			taupb, taunb, n1b, p1b, NAb=self.beforefitparams
			taupa, tauna, n1a, p1a, NAa=self.afterfitparams		
			nbefore=((n1b*taupb)+(taunb*(NAb+p1b))-((self.taubefore/1E6)*NAb))/((self.taubefore/1E6)-(taupb+taunb))
			nafter=((n1a*taupa)+(tauna*(NAa+p1a))-((self.tauafter/1E6)*NAa))/((self.tauafter/1E6)-(taupa+tauna))
			#nbefore[nbefore<0]=0
			#nafter[nafter<0]=0
			diffmatrix=np.abs(nafter-nbefore)
			self.diffmatrix=diffmatrix
			matrixmean= np.mean(diffmatrix[100:-100,100:-100])
			matrixstd= np.std(diffmatrix[100:-100,100:-100])			
			#if matrixmean>self.ddnlimit:
			diffmatrixflatten=diffmatrix.flatten()
			i=0
			count=0
			indicesb=[]
			indices=[]
			zeroarray=[]
			zeroarray2=[]
		# if self.canfit==1:
		# 	indices=[]
		# 	zeroarray=[]
		# 	oldshape=self.ironconcmatrixindividual.shape
		# 	ironflat=self.ironconcmatrixindividual.flatten()
		# 	i=0
		# 	while i<ironflat.size:
		# 		if ironflat[i]==np.Inf or ironflat[i]==-np.Inf or np.isnan(ironflat[i]):
		# 			indices.append(i)
		# 			zeroarray.append(-6e30)
		# 		i+=1

		# 	ironflat.put(indices, zeroarray)
		# 	self.ironconcmatrixindividual=ironflat.reshape(oldshape)
		# oldshape=self.ironconcmatrixindividual.shape
		# flatbefore=self.ironconcmatrixindividual.flatten()
		# #print flatbefore.size
		# flatbefore.put(indicesb, zeroarray2)
		# self.ironconcmatrixindividual=flatbefore.reshape(oldshape)

			nmean=(nafter+nbefore)/2.0
			sum=0
			count=0
			C=concentration.calcPrefactor(self.constants, float(self.builder.get_object("pldopingtxt").get_text()), nmean)
			self.ironconcmatrixindividual=np.abs(C*(1/self.taubefore - 1/self.tauafter))			
			#print diffmatrixflatten.size
			oldshape=self.ironconcmatrixindividual.shape
			ironflat=self.ironconcmatrixindividual.flatten()			
			while i<diffmatrixflatten.size:
				if diffmatrixflatten[i]>self.ddnlimit:
					indicesb.append(i)
					zeroarray2.append(-7E30)
					count+=1				
				elif ironflat[i]==np.Inf or ironflat[i]==-np.Inf or np.isnan(ironflat[i]):
					indices.append(i)
					zeroarray.append(-6E30)					
				i+=1
			#mean
			ironflat.put(indices, zeroarray)
			ironflat.put(indicesb, zeroarray2)
			self.ironconcmatrixindividual=ironflat.reshape(oldshape)
			if self.ironplottypebox.get_active()==1:			
				percent=(float(count)/diffmatrixflatten.size)*100
				self.builder.get_object("warnmeanlabel").set_label("The warning level is %.4g.\nThe number of values above this was %i out of %i pixels.\nThe percentage of values with differences above this was %.4g%%. \nThe mean of the array of differences of injection level values was %.4g.\nThe standard deviation of the array of differences of the injection level was %.4g\nThese values will be set to -7E30 on the PL map (and so coloured differently).\nThe warning level can be changed in the Edit Plot Options window.\nYou must recalculate the maps after changing this though." % (self.ddnlimit, count, diffmatrixflatten.size, percent, matrixmean, matrixstd))
				self.builder.get_object("warnpixlabel").set_label("%i/%i (%.4g%%)" % (count, diffmatrixflatten.size, percent))
				self.builder.get_object("validpointstxt").set_label("%i/%i (%.4g%%)" % (diffmatrixflatten.size-count, diffmatrixflatten.size, 100.0-percent))			
				self.builder.get_object("meanarraydifflabel").set_label("%.4g" % matrixmean)
				self.builder.get_object("stdarraydifflabel").set_label("%.4g" % matrixstd)
				self.builder.get_object("qsspcmeanwarning").show()
								 




		

		a1 = a[500]
		b1 = b[500]
		e1 = e[500]
		a1.sort()
		b1.sort()
		e1.sort()
		#Must make 0 tau values red
		self.taubmin=0
		self.tauamin=0
		self.taubmax = b1[int(0.99*len(b))]
		self.tauamax = a1[int(0.99*len(a))]
		self.ironmin=1E8
		self.ironmax = e1[int(0.97*len(e))]
		


		indices=[]
		zeroarray=[]
		oldshape=self.ironconcmatrixgenlevel.shape
		ironflat=self.ironconcmatrixgenlevel.flatten()
		i=0
		while i<ironflat.size:
			if ironflat[i]==np.Inf or ironflat[i]==-np.Inf or np.isnan(ironflat[i]):
				indices.append(i)
				zeroarray.append(-6e30)
			i+=1

		ironflat.put(indices, zeroarray)
		self.ironconcmatrixgenlevel=ironflat.reshape(oldshape)

		indices=[]
		zeroarray=[]
		oldshape=self.ironconcmatrixmeanC.shape
		ironflat=self.ironconcmatrixmeanC.flatten()
		i=0
		while i<ironflat.size:
			if ironflat[i]==np.Inf or ironflat[i]==-np.Inf or np.isnan(ironflat[i]):
				indices.append(i)
				zeroarray.append(-6e30)
			i+=1

		ironflat.put(indices, zeroarray)
		self.ironconcmatrixmeanC=ironflat.reshape(oldshape)

		self.builder.get_object("lifemintxt").set_text("%.4g" %self.taubmin)
		self.builder.get_object("lifemaxtxt").set_text("%.4g" %self.taubmax)
		self.builder.get_object("femintxt").set_text("%.4g" %self.ironmin)
		self.builder.get_object("femaxtxt").set_text("%.4g" %self.ironmax)

		oldshape=self.taubefore.shape
		flatbefore=self.taubefore.flatten()
		flatbefore.put(self.indicesb, self.negarrayb)
		self.taubefore=flatbefore.reshape(oldshape)
		oldshape=self.tauafter.shape
		flatafter=self.tauafter.flatten()		
		flatafter.put(self.indicesa, zeroarray)
		self.tauafter=flatafter.reshape(oldshape)

		self.plplotallbtnclicked(widget)

		#print 	 np.min(self.ironconcmatrixindividual[100:-100,100:-100])
		#random numbers chosen away from edges to avoid infinities
		#if self.ironconcmatrixgenlevel[547:555].mean()<=0:
		#	self.builder.get_object("negiron").show()

		#Fill show iron value boxes
		#1. QSSPC iron value from given generation/injection level: must interpolate iron plot
		self.ironcalcbtnclicked(widget)
		try:
			ironinterpf=scipy.interpolate.interp1d(self.qsspclist[0], self.qsspclist[1])
			iron=ironinterpf(float(self.builder.get_object("injectionleveltxt").get_text()))
			self.builder.get_object("qsspcgenlevel").set_text("%.4g" % iron)
		except: 
			self.builder.get_object("qsspcgenlevel").set_text("Given Injection level out of bounds for interpolation!")
		self.builder.get_object("lastqsspciron").set_text("%.4g" % self.qsspcironmeanlast)
		self.builder.get_object("stdlastqsspciron").set_text("%.4g" % self.qsspcstdev)
		#mean gen and tau levels
		cutborders=self.ironconcmatrixindgen[100:-100,100:-100]
		i=0
		j=0
		n=0
		total=0
		while i<800:
			while j<800:
				if cutborders[i][j]!=np.Inf and cutborders[i][j]!=-np.Inf and cutborders[i][j]!=np.NaN and cutborders[i][j]>0:
					total+=cutborders[i][j]
					n+=1
				j+=1
			j=0
			i+=1
		mean=total/float(n)
		self.builder.get_object("meanironplindgen").set_text("%.4g" % mean)
		i=0
		j=0
		n=0
		total=0
		while i<800:
			while j<800:
				if cutborders[i][j]!=np.Inf and cutborders[i][j]!=-np.Inf and cutborders[i][j]!=np.NaN and cutborders[i][j]>0:
					total+=pow((cutborders[i][j]-mean),2)
					n+=1
				j+=1
			j=0
			i+=1
		#print total
		#print n
		stddev=np.sqrt((1.0/(n-1))*total)
		self.builder.get_object("stdironplindgen").set_text("%.4g" % stddev)
		
		if self.canfit==1:
			if np.max(self.ironconcmatrixindividual[100:-100,100:-100]) != np.Inf and np.min(self.ironconcmatrixindividual[100:-100,100:-100]) !=-np.Inf and np.min(self.ironconcmatrixindividual[100:-100,100:-100])>0:
				self.builder.get_object("meanironplindividual").set_text("%.4g" % np.mean(self.ironconcmatrixindividual[100:-100,100:-100]))
				self.builder.get_object("stdironplindividual").set_text("%.4g" % np.std(self.ironconcmatrixindividual[100:-100,100:-100]))
			else:
				cutborders=self.ironconcmatrixindividual[100:-100,100:-100]
				i=0
				j=0
				n=0
				total=0
				while i<800:
					while j<800:
						if cutborders[i][j]!=np.Inf and cutborders[i][j]!=-np.Inf and cutborders[i][j]!=np.NaN and cutborders[i][j]>0:
							total+=cutborders[i][j]
							n+=1
						j+=1
					j=0
					i+=1
				mean=total/float(n)
				self.builder.get_object("meanironplindividual").set_text("%.4g" % mean)
				i=0
				j=0
				n=0
				total=0
				while i<800:
					while j<800:
						if cutborders[i][j]!=np.Inf and cutborders[i][j]!=-np.Inf and cutborders[i][j]!=np.NaN and cutborders[i][j]>0:
							total+=pow((cutborders[i][j]-mean),2)
							n+=1
						j+=1
					j=0
					i+=1
				#print total
				#print n
				stddev=np.sqrt((1.0/(n-1))*total)
				self.builder.get_object("stdironplindividual").set_text("%.4g" % stddev)
				
		else:
			self.builder.get_object("meanironplindividual").set_text("Fit could not be calculated for data!")
		
		if np.max(self.ironconcmatrixgenlevel[100:-100,100:-100]) != np.Inf and np.min(self.ironconcmatrixgenlevel[100:-100,100:-100]) !=-np.Inf and np.min(self.ironconcmatrixgenlevel[100:-100,100:-100])>0:
			self.builder.get_object("meanironplgenlevel").set_text("%.4g" % np.mean(self.ironconcmatrixgenlevel[100:-100,100:-100]))
			self.builder.get_object("stdironplgenlevel").set_text("%.4g" % np.std(self.ironconcmatrixgenlevel[100:-100,100:-100]))			
		else:
			cutborders=self.ironconcmatrixgenlevel[100:-100,100:-100]
			i=0
			j=0
			n=0
			total=0
			while i<800:
				while j<800:
					if cutborders[i][j]!=np.Inf and cutborders[i][j]!=-np.Inf and cutborders[i][j]!=np.NaN and cutborders[i][j]>0:
						total+=cutborders[i][j]
						n+=1
					j+=1
				j=0
				i+=1
			mean=total/float(n)
			self.builder.get_object("meanironplgenlevel").set_text("%.4g" % mean)
			i=0
			j=0
			n=0
			total=0
			while i<800:
				while j<800:
					if cutborders[i][j]!=np.Inf and cutborders[i][j]!=-np.Inf and cutborders[i][j]!=np.NaN and cutborders[i][j]>0:
						total+=pow((cutborders[i][j]-mean),2)
						n+=1
					j+=1
				j=0
				i+=1
			stddev=np.sqrt((1.0/(n-1))*total)
			self.builder.get_object("stdironplgenlevel").set_text("%.4g" % stddev)

		if np.max(self.ironconcmatrixmeanC[100:-100,100:-100]) != np.Inf and np.min(self.ironconcmatrixmeanC[100:-100,100:-100]) !=-np.Inf and np.min(self.ironconcmatrixmeanC[100:-100,100:-100])>0 :
			self.builder.get_object("meanironpllastc").set_text("%.4g" % np.mean(self.ironconcmatrixmeanC[100:-100,100:-100]))
			self.builder.get_object("stdironpllastc").set_text("%.4g" % np.std(self.ironconcmatrixmeanC[100:-100,100:-100]))			
		else:
			cutborders=self.ironconcmatrixmeanC[100:-100,100:-100]
			i=0
			j=0
			n=0
			total=0
			while i<800:
				while j<800:
					if cutborders[i][j]!=np.Inf and cutborders[i][j]!=-np.Inf and cutborders[i][j]!=np.NaN and cutborders[i][j]>0:
						total+=cutborders[i][j]
						n+=1
					j+=1
				j=0
				i+=1
			mean=total/float(n)
			self.builder.get_object("meanironpllastc").set_text("%.4g" % mean)
			i=0
			j=0
			n=0
			total=0
			while i<800:
				while j<800:
					if cutborders[i][j]!=np.Inf and cutborders[i][j]!=-np.Inf and cutborders[i][j]!=np.NaN and cutborders[i][j]>0:
						total+=pow((cutborders[i][j]-mean),2)
						n+=1
					j+=1
				j=0
				i+=1
			stddev=np.sqrt((1.0/(n-1))*total)
			self.builder.get_object("stdironpllastc").set_text("%.4g" % stddev)			

		self.builder.get_object("showironvaluesbtn").set_sensitive(True)

		#except:
			#self.errorshow(widget)

		if self.adjusted==0:
		#cross-correlation - ignored until can fix Mirco's program
#			self.oldcc=((self.tauafter-self.tauafter.mean()) * (self.taubefore - self.taubefore.mean())).mean() / (self.taubefore.std() * self.tauafter.std())
			self.oldcc=bildreg.correlation(self.taubefore[100:-100,100:-100], self.tauafter[100:-100,100:-100])
			#if cc<0.95:
			#	self.builder.get_object("correlationwarning").set_property("secondary-text", "There was a significant difference between the loaded PL images, this means there may have been a change in the position of the sample between the taking of the images.\nCross-correlation value: %.4g (anything below 0.95 is considered anomalous)" % cc)
			#	self.builder.get_object("correlationwarning").show()

			self.builder.get_object("correlationlabel").set_label("The Cross-Correlation value for the images is %.4g\n <b>Should image matching be attempted?</b>\n(note this only guarantees an improvement below 0.85, at high correlations it may produce a worse image!)\n" % self.oldcc) 
			self.builder.get_object("imagematchingwindow").show()

			self.builder.get_object("compareplmapsbtn").set_sensitive(True)

	def attemptcorrectionbtnclicked(self, widget):
		self.builder.get_object("progressbar1").set_fraction(0)
		self.builder.get_object("progressbar1").set_visible(True)
		trafo = bildreg.transformation_berechnen(self.taubefore, self.tauafter, maxdelta=34, korrelationsgroesse=128, N=5, callback=self.progbar)
		T_T = np.ndarray(self.tauafter.shape, dtype=np.float)
		bildreg.transform(self.tauafter,trafo,T_T)
		self.tauafter=T_T
		self.adjusted=1

		self.plcalcbtnclicked(widget)
		self.newcc=bildreg.correlation(self.taubefore[100:-100,100:-100], self.tauafter[100:-100,100:-100])
#((self.tauafter-self.tauafter.mean()) * (self.taubefore - self.taubefore.mean())).mean() / (self.taubefore.std() * self.tauafter.std())
		self.plcalcbtnclicked(widget)
		self.builder.get_object("progressbar1").set_visible(False)
		self.builder.get_object("imagematchingwindow").hide()
		self.builder.get_object("matchingcomplete").set_property("secondary-text", "The lifetime map after illumination was translated on to the lifetime map before illumination. The new cross-correlation value is %.4g, compared to the old one of %.4g" % (self.newcc, self.oldcc))
		self.builder.get_object("matchingcomplete").show()

	def progbar(self, number):
		self.builder.get_object("progressbar1").set_fraction(number)
		gtk.main_iteration()
		
	def donothingbtnclicked(self, widget):
		self.builder.get_object("imagematchingwindow").hide()
		
	def plplotallbtnclicked(self, widget):

		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]

		if self.plcounter==0:
			self.figure3 = Figure(figsize=(6,4), dpi=72)  
			self.figure4=Figure(figsize=(6,4), dpi=72)
		if self.plcounter>0:
			self.axis3.clear()
			self.axis4.clear()
			self.axis5.clear()
			self.figure3.clear()
			self.figure4.clear()
			self.builder.get_object("plgraphs1").remove(self.canvas3)
			self.builder.get_object("plgraphs2").remove(self.canvas4)
			l_f = LogFormatter(10, labelOnlyBase=False)
		if self.lifemaptype=="logarithmic":
			self.axis3 = self.figure3.add_subplot(211, aspect='equal')
			imagebefore=self.axis3.imshow(self.taubefore, cmap=self.cmap, norm=LogNorm(vmin=self.taubmin, vmax=self.taubmax))
			cbar1 = self.figure3.colorbar(imagebefore, format = l_f, extend='both') 
			self.axis3.set_title("Lifetime before illumination")
			self.axis4=self.figure3.add_subplot(212, aspect='equal')
			imageafter=self.axis4.imshow(self.tauafter, cmap=self.cmap, norm=LogNorm(vmin=self.taubmin, vmax=self.taubmax))
			cbar2=self.figure3.colorbar(imageafter, format=l_f, extend='both')

		elif self.lifemaptype=="linear":
			self.axis3 = self.figure3.add_subplot(211, aspect='equal')
			imagebefore=self.axis3.imshow(self.taubefore, cmap=self.cmap,vmin=self.taubmin, vmax=self.taubmax)
			cbar1=self.figure3.colorbar(imagebefore, orientation="vertical", extend='both')
			self.axis3.set_title("Lifetime before illumination")
			self.axis4=self.figure3.add_subplot(212, aspect='equal')
			imageafter=self.axis4.imshow(self.tauafter, cmap=self.cmap, vmin=self.taubmin, vmax=self.taubmax)
			cbar2=self.figure3.colorbar(imageafter, extend='both')

		cbar1.set_label("Lifetime ($\mu$s)", rotation='vertical')
		cbar2.ax.set_ylabel("Lifetime ($\mu$s)", rotation='vertical', verticalalignment='center') #rotation doesnt work with mathtex
		self.axis4.set_title("Lifetime after illumination")
		self.canvas3=FigureCanvasGTKAgg(self.figure3)
		self.canvas3.show()

		self.axis5=self.figure4.add_subplot(111, aspect='equal')
		if self.ironmaptype=="linear":
			if self.ironplottypebox.get_active()==0:
				#Individual C values from Gen level, tau values
				imageiron=self.axis5.imshow(self.ironconcmatrixindgen, cmap=self.cmap, vmin=self.ironmin, vmax=self.ironmax)
			if self.ironplottypebox.get_active()==1:
				#Individual C values
				imageiron=self.axis5.imshow(self.ironconcmatrixindividual, cmap=self.cmap, vmin=self.ironmin, vmax=self.ironmax)
			elif self.ironplottypebox.get_active()==2:
				#C value from Generation Level
				imageiron=self.axis5.imshow(self.ironconcmatrixgenlevel, cmap=self.cmap, vmin=self.ironmin, vmax=self.ironmax)
			elif self.ironplottypebox.get_active()==3:
				#Use mean of last C values
				imageiron=self.axis5.imshow(self.ironconcmatrixmeanC, cmap=self.cmap, vmin=self.ironmin, vmax=self.ironmax)

			cbar3=self.figure4.colorbar(imageiron, fraction=0.045, extend='both')
		elif self.ironmaptype=="logarithmic":
			if self.ironplottypebox.get_active()==0:
				#Individual C values from Gen level, tau values				
				imageiron=self.axis5.imshow(self.ironconcmatrixindgen, cmap=self.cmap, norm=LogNorm(vmin=self.ironmin, vmax=self.ironmax))			
			if self.ironplottypebox.get_active()==1:
				#Individual C values
				imageiron=self.axis5.imshow(self.ironconcmatrixindividual, cmap=self.cmap, norm=LogNorm(vmin=self.ironmin, vmax=self.ironmax))
			elif self.ironplottypebox.get_active()==2:
				#C value from Generation Level
				imageiron=self.axis5.imshow(self.ironconcmatrixgenlevel, cmap=self.cmap, norm=LogNorm(vmin=self.ironmin, vmax=self.ironmax))
			elif self.ironplottypebox.get_active()==3:
				imageiron=self.axis5.imshow(self.ironconcmatrixmeanC, cmap=self.cmap, norm=LogNorm(vmin=self.ironmin, vmax=self.ironmax))
				#Use mean of last C values
			cbar3=self.figure4.colorbar(imageiron, format=l_f, fraction=0.045, extend='both')

		cbar3.set_label(r'[Fe$_{i}$]  (cm$^{-3})$',fontsize=16)
		self.axis5.set_title("Interstitial Iron concentration", fontsize=14)
		self.axis5.set_xlabel("Pixels", fontsize=14)
		self.axis5.set_ylabel("Pixels", fontsize=14)
		self.axis4.set_xlabel("Pixels", fontsize=14)
		self.axis4.set_ylabel("Pixels", fontsize=14)
		self.axis3.set_xlabel("Pixels", fontsize=14)
		self.axis3.set_ylabel("Pixels", fontsize=14)			
		self.canvas4=FigureCanvasGTKAgg(self.figure4)
		self.canvas4.show()
		
		self.axis3.set_xlim(left=0, right=1000)
		self.axis3.set_ylim(top=0, bottom=1000)
		self.axis4.set_xlim(left=0, right=1000)
		self.axis4.set_ylim(top=0, bottom=1000)
		self.axis5.set_xlim(left=0, right=1000)
		self.axis5.set_ylim(top=0, bottom=1000)
		
		self.builder.get_object("plgraphs1").pack_start(self.canvas3, True, True)
		self.builder.get_object("plgraphs2").pack_start(self.canvas4, True, True)
		self.builder.get_object("pltoolbar").set_sensitive(True)
		self.builder.get_object("editcolorbarbtn").set_sensitive(True)
		self.plcounter+=1

		self.canvas4.mpl_connect('motion_notify_event', self.graphscroll)
		self.canvas3.mpl_connect('motion_notify_event', self.graphscroll)

	def zoominbtnclicked(self, widget):
		xlims=self.axis5.get_xlim()
		ylims=self.axis5.get_ylim()
		xchange=abs(xlims[1]-xlims[0])*0.1
		ychange=abs(ylims[1]-ylims[0])*0.1
		self.axis3.set_xlim(left=xlims[0]+xchange, right=xlims[1]-xchange)
		self.axis3.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]-ychange)
		self.axis4.set_xlim(left=xlims[0]+xchange, right=xlims[1]-xchange)
		self.axis4.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]-ychange)
		self.axis5.set_xlim(left=xlims[0]+xchange, right=xlims[1]-xchange)
		self.axis5.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]-ychange)
		self.builder.get_object("plgraphs1").remove(self.canvas3)
		self.builder.get_object("plgraphs2").remove(self.canvas4)
		self.builder.get_object("plgraphs1").pack_start(self.canvas3, True, True)
		self.builder.get_object("plgraphs2").pack_start(self.canvas4, True, True)
		
	def zoomoutbtnclicked(self, widget):
		xlims=self.axis5.get_xlim()
		ylims=self.axis5.get_ylim()
		xchange=abs(xlims[1]-xlims[0])*0.1
		xchange=xchange/0.8
		ychange=abs(ylims[1]-ylims[0])*0.1
		ychange=ychange/0.8
		self.axis3.set_xlim(left=xlims[0]-xchange, right=xlims[1]+xchange)
		self.axis3.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]+ychange)
		self.axis4.set_xlim(left=xlims[0]-xchange, right=xlims[1]+xchange)
		self.axis4.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]+ychange)
		self.axis5.set_xlim(left=xlims[0]-xchange, right=xlims[1]+xchange)
		self.axis5.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]+ychange)
		self.builder.get_object("plgraphs1").remove(self.canvas3)
		self.builder.get_object("plgraphs2").remove(self.canvas4)
		self.builder.get_object("plgraphs1").pack_start(self.canvas3, True, True)
		self.builder.get_object("plgraphs2").pack_start(self.canvas4, True, True)

       	def panleftbtnclicked(self, widget):
		xlims=self.axis5.get_xlim()
		xchange=abs(xlims[1]-xlims[0])*0.1
		self.axis3.set_xlim(left=xlims[0]-xchange, right=xlims[1]-xchange)
		self.axis4.set_xlim(left=xlims[0]-xchange, right=xlims[1]-xchange)
		self.axis5.set_xlim(left=xlims[0]-xchange, right=xlims[1]-xchange)
		self.builder.get_object("plgraphs1").remove(self.canvas3)
		self.builder.get_object("plgraphs2").remove(self.canvas4)
		self.builder.get_object("plgraphs1").pack_start(self.canvas3, True, True)
		self.builder.get_object("plgraphs2").pack_start(self.canvas4, True, True)

       	def panrightbtnclicked(self, widget):
		xlims=self.axis5.get_xlim()
		xchange=abs(xlims[1]-xlims[0])*0.1
		self.axis3.set_xlim(left=xlims[0]+xchange, right=xlims[1]+xchange)
		self.axis4.set_xlim(left=xlims[0]+xchange, right=xlims[1]+xchange)
		self.axis5.set_xlim(left=xlims[0]+xchange, right=xlims[1]+xchange)
		self.builder.get_object("plgraphs1").remove(self.canvas3)
		self.builder.get_object("plgraphs2").remove(self.canvas4)
		self.builder.get_object("plgraphs1").pack_start(self.canvas3, True, True)
		self.builder.get_object("plgraphs2").pack_start(self.canvas4, True, True)

	def panupbtnclicked(self, widget):
		ylims=self.axis5.get_ylim()
		ychange=abs(ylims[1]-ylims[0])*0.1
		self.axis3.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]-ychange)
		self.axis4.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]-ychange)
		self.axis5.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]-ychange)
		self.builder.get_object("plgraphs1").remove(self.canvas3)
		self.builder.get_object("plgraphs2").remove(self.canvas4)
		self.builder.get_object("plgraphs1").pack_start(self.canvas3, True, True)
		self.builder.get_object("plgraphs2").pack_start(self.canvas4, True, True)

	def pandownbtnclicked(self, widget):
		ylims=self.axis5.get_ylim()
		ychange=abs(ylims[1]-ylims[0])*0.1
		self.axis3.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]+ychange)
		self.axis4.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]+ychange)
		self.axis5.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]+ychange)
		self.builder.get_object("plgraphs1").remove(self.canvas3)
		self.builder.get_object("plgraphs2").remove(self.canvas4)
		self.builder.get_object("plgraphs1").pack_start(self.canvas3, True, True)
		self.builder.get_object("plgraphs2").pack_start(self.canvas4, True, True)
	
	def savefemapbtnclicked(self, widget):
		self.builder.get_object("whichsavewindow").show()

	def whichsavefebtnclicked(self, widget):
		self.cursavepl="femap"
		self.builder.get_object("whichsavewindow").hide()
		self.builder.get_object("plplotfilesavedialog").set_property("title", "Save interstitial Iron concentration map to PNG")
		self.builder.get_object("plplotfilesavedialog").show()

	def whichsavelifebtnclicked(self, widget):
		self.cursavepl="lifemap"
		self.builder.get_object("whichsavewindow").hide()
		self.builder.get_object("plplotfilesavedialog").set_property("title", "Save lifetime maps to PNG")
		self.builder.get_object("plplotfilesavedialog").show()


	def whichsavecancelbtnclicked(self, widget):
		self.builder.get_object("whichsavewindow").hide()		

	def plplotdialogsavebtnclicked(self, widget):
		#atm only does iron map
		self.issaving=True
		self.cursave="pl"
		filename=self.builder.get_object("plplotfilesavedialog").get_filename()
		if filename[-4:len(filename)]!=".png" and filename[-4:len(filename)]!=".PNG":
			filename += ".png"
		self.currentfilename=filename
		if os.path.exists(filename)==True:
			self.builder.get_object("overwritewarning").show()
		else:
			self.saveplmap()

			
	
	def plplotdialogcancelbtnclicked(self, widget):
		self.builder.get_object("plplotfilesavedialog").hide()
		
	def overwriteresponseclicked(self, widget, response):
		#yes -8 no -9
		self.issaving=True
		self.builder.get_object("overwritewarning").hide()
		if response==-8:
			if self.cursave=="pl":
				self.saveplmap()
			elif self.cursave=="plotdialogsave":
				self.figure2.savefig(self.currentfilename, format="png")
				self.builder.get_object("plotfilesavedialog").hide()	
			elif self.cursave=="feplotdata":
				self.datatxtsave()
			elif self.cursave=="femapdata":
				if self.ironplottypebox.get_active()==0:
				#Individual C values from gen level, tau values
					savetxt(self.currentfilename, self.ironconcmatrixindgen, fmt="%12.6G")				
				if self.ironplottypebox.get_active()==1:
				#Individual C values from fits
					savetxt(self.currentfilename, self.ironconcmatrixindividual, fmt="%12.6G")
				elif self.ironplottypebox.get_active()==2:
				#C value from Generation Level
					savetxt(self.currentfilename, self.ironconcmatrixgenlevel, fmt="%12.6G")
				elif self.ironplottypebox.get_active()==3:
					savetxt(self.currentfilename, self.ironconcmatrixmeanC, fmt="%12.6G")
				#Use mean of last C values
			elif self.cursave=="adjusteddata":
				savetxt(self.currentfilename, self.tauafter, fmt="%12.6G")
	def saveplmap(self):
		self.builder.get_object("plplotfilesavedialog").hide()
		#add error handling
		#check if filename has extension, if not append PNG - maybe ask?
		l_f = LogFormatter(10, labelOnlyBase=False)
		if self.cursavepl=="femap":
			xlims=self.axis5.get_xlim()
			ylims=self.axis5.get_ylim()
			if self.issaving==True:
				self.figure4.savefig(self.currentfilename, format="png")
			self.axis5.clear()
			self.figure4.clear()
			self.builder.get_object("plgraphs2").remove(self.canvas4)
			self.axis5=self.figure4.add_subplot(111, aspect='equal')

			if self.ironmaptype=="linear":
				if self.ironplottypebox.get_active()==0:
				#Individual C values from gen level, tau values
					imageiron=self.axis5.imshow(self.ironconcmatrixindgen, cmap=self.cmap, vmin=self.ironmin, vmax=self.ironmax)				
				if self.ironplottypebox.get_active()==1:
				#Individual C values
					imageiron=self.axis5.imshow(self.ironconcmatrixindividual, cmap=self.cmap, vmin=self.ironmin, vmax=self.ironmax)
				elif self.ironplottypebox.get_active()==2:
				#C value from Generation Level
					imageiron=self.axis5.imshow(self.ironconcmatrixgenlevel, cmap=self.cmap, vmin=self.ironmin, vmax=self.ironmax)
				elif self.ironplottypebox.get_active()==3:
				#Use mean of last C values
					imageiron=self.axis5.imshow(self.ironconcmatrixmeanC, cmap=self.cmap, vmin=self.ironmin, vmax=self.ironmax)

				cbar3=self.figure4.colorbar(imageiron, fraction=0.045, extend='both')
			elif self.ironmaptype=="logarithmic":
				if self.ironplottypebox.get_active()==0:
				#Individual C values from tau values, gen level
					imageiron=self.axis5.imshow(self.ironconcmatrixindgen, cmap=self.cmap, norm=LogNorm(vmin=self.ironmin, vmax=self.ironmax))				
				if self.ironplottypebox.get_active()==1:
				#Individual C values
					imageiron=self.axis5.imshow(self.ironconcmatrixindividual, cmap=self.cmap, norm=LogNorm(vmin=self.ironmin, vmax=self.ironmax))
				elif self.ironplottypebox.get_active()==2:
				#C value from Generation Level
					imageiron=self.axis5.imshow(self.ironconcmatrixgenlevel, cmap=self.cmap, norm=LogNorm(vmin=self.ironmin, vmax=self.ironmax))
				elif self.ironplottypebox.get_active()==3:
					imageiron=self.axis5.imshow(self.ironconcmatrixmeanC, cmap=self.cmap, norm=LogNorm(vmin=self.ironmin, vmax=self.ironmax))
				cbar3=self.figure4.colorbar(imageiron, fraction=0.045, format=l_f, extend='both')

			cbar3.set_label(r'[Fe$_{i}$]  (cm$^{-3})$', fontsize=16)
			self.axis5.set_title("Interstitial Iron concentration", fontsize=14)
			self.axis5.set_xlabel("Pixels", fontsize=14)
			self.axis5.set_ylabel("Pixels", fontsize=14)			
			self.canvas4=FigureCanvasGTKAgg(self.figure4)
			self.canvas4.show()

			self.axis5.set_xlim(xlims)
			self.axis5.set_ylim(ylims)

			self.builder.get_object("plgraphs2").pack_start(self.canvas4, True, True)
			self.canvas4.mpl_connect('motion_notify_event', self.graphscroll)

		elif self.cursavepl=="lifemap":
			xlims=self.axis5.get_xlim()
			ylims=self.axis5.get_ylim()
			if self.issaving==True:
				self.figure3.savefig(self.currentfilename, format="png")

			self.axis3.clear()
			self.axis4.clear()
			self.figure3.clear()
			self.builder.get_object("plgraphs1").remove(self.canvas3)
			
			if self.lifemaptype=="logarithmic":
				self.axis3 = self.figure3.add_subplot(211, aspect='equal')
				imagebefore=self.axis3.imshow(self.taubefore, cmap=self.cmap, norm=LogNorm(vmin=self.taubmin, vmax=self.taubmax))
				cbar1 = self.figure3.colorbar(imagebefore, format = l_f, extend='both') 
				self.axis3.set_title("Lifetime before illumination")
				self.axis4=self.figure3.add_subplot(212, aspect='equal')
				imageafter=self.axis4.imshow(self.tauafter, cmap=self.cmap, norm=LogNorm(vmin=self.taubmin, vmax=self.taubmax))
				cbar2=self.figure3.colorbar(imageafter, format=l_f, extend='both')

			elif self.lifemaptype=="linear":
				self.axis3 = self.figure3.add_subplot(211, aspect='equal')
				imagebefore=self.axis3.imshow(self.taubefore, cmap=self.cmap,vmin=self.taubmin, vmax=self.taubmax)
				cbar1=self.figure3.colorbar(imagebefore, orientation="vertical", extend='both')
				self.axis3.set_title("Lifetime before illumination")
				self.axis4=self.figure3.add_subplot(212, aspect='equal')
				imageafter=self.axis4.imshow(self.tauafter, cmap=self.cmap, vmin=self.taubmin, vmax=self.taubmax)
				cbar2=self.figure3.colorbar(imageafter, extend='both')


			cbar1.set_label("Lifetime ($\mu$s)", rotation='vertical', verticalalignment='center')
			cbar2.set_label("Lifetime ($\mu$s)")
			self.axis4.set_title("Lifetime after illumination")
			self.canvas3=FigureCanvasGTKAgg(self.figure3)
			self.canvas3.show()

			self.axis3.set_xlim(xlims)
			self.axis3.set_ylim(ylims)
			self.axis4.set_xlim(xlims)
			self.axis4.set_ylim(ylims)
			self.axis4.set_xlabel("Pixels", fontsize=14)
			self.axis4.set_ylabel("Pixels", fontsize=14)
			self.axis3.set_xlabel("Pixels", fontsize=14)
			self.axis3.set_ylabel("Pixels", fontsize=14)			
			self.canvas3.mpl_connect('motion_notify_event', self.graphscroll)

			self.builder.get_object("plgraphs1").pack_start(self.canvas3, True, True)

	def pl2genlevelokbtnclicked(self, widget):
		try:
			self.builder.get_object("pl2genleveldialog").hide()
			genlevelaim=float(self.builder.get_object("pl2pfluxb").get_text())*(1.0-float(self.builder.get_object("pl2reflectivity").get_text()))
			self.genlevelbefore=float(self.builder.get_object("pl2pfluxb").get_text())*(1.0-float(self.builder.get_object("pl2reflectivity").get_text()))
			self.genlevelafter=float(self.builder.get_object("pl2pfluxa").get_text())*(1.0-float(self.builder.get_object("pl2reflectivity").get_text()))
			filename1=self.builder.get_object("filebeforebtn").get_filename()
			closestlevel=0
			closesti=0
			wb1 = load_workbook(filename1)
			ws1 = wb1.get_sheet_by_name("Calc")
			i=15
			for row in ws1.range('K15:K131'):
				for cell in row:
					curdiff=abs(genlevelaim-float(cell.value))
					if curdiff<abs(genlevelaim-closestlevel):
						closestlevel=float(cell.value)
						closesti=i

					i+=1

			injectionlevelaim=float(ws1.cell('S'+str(closesti)).value)
			self.injectionlevel=injectionlevelaim
			#get wanted generation level for second file
			closestlevel=0
			closesti=0
			filename2=self.builder.get_object("fileafterbtn").get_filename()
			wb2 = load_workbook(filename2)
			ws2 = wb2.get_sheet_by_name("Calc")
			i=15
			for row in ws2.range('S15:S131'):
				for cell in row:
					curdiff=abs(injectionlevelaim-float(cell.value))
					if curdiff<abs(injectionlevelaim-closestlevel):
						closestlevel=float(cell.value)
						closesti=i

					i+=1

			wantedilevel=float(ws2.cell('K'+str(closesti)).value)
			#print float(ws2.cell('S'+str(closesti)).value)
			self.builder.get_object("plgenleveltxt").set_text("%.4g" % wantedilevel)
			self.builder.get_object("injectionleveltxt").set_text("%.4g" % injectionlevelaim)
			self.builder.get_object("plafterfile").set_property("title", "Open PL data after illumination, G~%.3g" % wantedilevel)

			self.builder.get_object("plafterfile").set_sensitive(True)
			
		except:
			#change this to a proper error message later
			self.builder.get_object("ErrorWindow2").show()
			return 1

	def pl2genlevelcancelbtnclicked(self,widget):
		self.builder.get_object("pl2genleveldialog").hide()

	def getfebtnclicked(self, widget):
		
		if self.curplot=="Fe":
			#limit=self.irondata[1].index(max(self.irondata[1]))
			ironvalues=self.irondata[1]
			deltanvalues=self.irondata[0]
			interpf=interp1d(deltanvalues, ironvalues)
			try: 
				interpval=interpf(float(self.builder.get_object("deltangettxt").get_text()))
				#interpval=interp(float(self.builder.get_object("deltangettxt").get_text()), deltanvalues, ironvalues, left=-99.0, right=-99.0)
				#add error handling code
				self.builder.get_object("fegettxt").set_text("%.3e" % interpval)


			except:
				self.builder.get_object("interperrorwindow").show()
		elif self.curplot=="C":
			deltanvalues=self.cvaluedn
			interpf=interp1d(deltanvalues, self.cvalues)
			try: 
				interpval=interpf(float(self.builder.get_object("deltangettxt").get_text()))
				#interpval=interp(float(self.builder.get_object("deltangettxt").get_text()), deltanvalues, self.cvalues, left=-99.0, right=-99.0)
				self.builder.get_object("fegettxt").set_text("%.3e" % interpval)


			except:
				self.builder.get_object("interperrorwindow").show()
			
	def editcolorbarbtnclicked(self, widget):
		self.builder.get_object("ddnwarnlevel").set_text("%.4g" % self.ddnlimit)
	        self.builder.get_object("lifemintxt").set_text("%.4g" %self.taubmin)
		self.builder.get_object("lifemaxtxt").set_text("%.4g" %self.taubmax)
		self.builder.get_object("femintxt").set_text("%.4g" %self.ironmin)
		self.builder.get_object("femaxtxt").set_text("%.4g" %self.ironmax)
		self.oldirontypeselection=self.ironplottypebox.get_active()
		if self.lifemaptype=="linear":
			self.linbutton.set_active(True)
		if self.lifemaptype=="logarithmic":
			self.logbutton.set_active(True)
		if self.ironmaptype=="linear":
			self.felinbutton.set_active(True)
		if self.ironmaptype=="logarithmic":
			self.felogbutton.set_active(True)
		if self.cmap==cm.gray:
			self.combobox.set_active(0)
		elif self.cmap==cm.jet:
			self.combobox.set_active(1)
		elif self.cmap==cm.hot:
			self.combobox.set_active(2)
		self.builder.get_object("editcolorbarwindow").show()

	def editcolorbarcancelbtnclicked(self,widget):
		self.ironplottypebox.set_active(self.oldirontypeselection)
		self.builder.get_object("editcolorbarwindow").hide()

	def editcolorbarokbtnclicked(self,widget):
		self.ddnlimit=float(self.builder.get_object("ddnwarnlevel").get_text())
		if self.ironplottypebox.get_active()==0:
				#Individual C values
			self.builder.get_object("maptypelabel").set_label("<b>Individual C values from generation levels and tau values</b>")		
		if self.ironplottypebox.get_active()==1:
				#Individual C values
			self.builder.get_object("maptypelabel").set_label("<b>Individual C values from QSSPC fits</b>")
		elif self.ironplottypebox.get_active()==2:
				#C value from Generation Level
			self.builder.get_object("maptypelabel").set_label("<b>C value from Generation Level</b>")

		elif self.ironplottypebox.get_active()==3:
			self.builder.get_object("maptypelabel").set_label("<b>Mean C value from given range</b>")

		self.builder.get_object("editcolorbarwindow").hide()
	        self.taubmin=float(self.builder.get_object("lifemintxt").get_text())
	        self.taubmax=float(self.builder.get_object("lifemaxtxt").get_text())
	        self.ironmin=float(self.builder.get_object("femintxt").get_text())
	        self.ironmax=float(self.builder.get_object("femaxtxt").get_text())
		if self.combobox.get_active()==0:
			self.cmap=cm.gray
		elif self.combobox.get_active()==1:
			self.cmap=cm.jet
		elif self.combobox.get_active()==2:
			self.cmap=cm.hot
		self.ironmaptype=self.tempironmaptype
		self.lifemaptype=self.templifemaptype
		self.issaving=False
		self.cursavepl="femap"
		self.saveplmap()
		self.cursavepl="lifemap"
		self.saveplmap()

	def saveirondatabtnclicked(self, widget):
		self.datasave="map"
		self.builder.get_object("whichsavewindow").hide()
		self.builder.get_object("datafilesavedialog").show()

	def buttontoggle(self, widget, data=None):
		if widget.get_active()==True:
			self.templifemaptype=data

	def febuttontoggle(self, widget, data=None):
		if widget.get_active()==True:
			self.tempironmaptype=data

	def saveadjusteddatabtnclicked(self, widget):
		self.datasave="adjusteddata"
		self.builder.get_object("datafilesavedialog").show()
		

	def graphscroll(self, event):
		if event.x!=None and event.y!=None and event.xdata!=None and event.ydata!=None:

			if self.curid!=None:
				self.builder.get_object("statusbar").remove_message(self.plconid, self.curid)
				
			self.curid=self.builder.get_object("statusbar").push(self.plconid, 'x=%d, y=%d, taub=%.3g, taua=%.3g, irongenlevelC=%.3g, ironindividualC=%.3g, ironlastmeanC=%.3g, ironindgen=%.3g'%(int(round(event.xdata)), int(round(event.ydata)), self.taubefore[int(round(event.ydata))][int(round(event.xdata))], self.tauafter[int(round(event.ydata))][int(round(event.xdata))], self.ironconcmatrixgenlevel[int(round(event.ydata))][int(round(event.xdata))], self.ironconcmatrixindividual[int(round(event.ydata))][int(round(event.xdata))], self.ironconcmatrixmeanC[int(round(event.ydata))][int(round(event.xdata))], self.ironconcmatrixindgen[int(round(event.ydata))][int(round(event.xdata))]))

	def showironvaluesbtnclicked(self, widget):
		self.builder.get_object("ironviewwindow").show()

	def ironviewclosebtnclicked(self, widget):
		self.builder.get_object("ironviewwindow").hide()

	def changefitparamsbtnclicked(self, widget):
		self.builder.get_object("taupb").set_text("%.4g" % self.startguessb[0])
		self.builder.get_object("taunb").set_text("%.4g" % self.startguessb[1])
		self.builder.get_object("n1b").set_text("%.4g" % self.startguessb[2])
		self.builder.get_object("p1b").set_text("%.4g" % self.startguessb[3])
		self.builder.get_object("NAb").set_text("%.4g" % self.startguessb[4])
		self.builder.get_object("taupa").set_text("%.4g" % self.startguessa[0])
		self.builder.get_object("tauna").set_text("%.4g" % self.startguessa[1])
		self.builder.get_object("n1a").set_text("%.4g" % self.startguessa[2])
		self.builder.get_object("p1a").set_text("%.4g" % self.startguessa[3])
		self.builder.get_object("NAa").set_text("%.4g" % self.startguessa[4])
		self.builder.get_object("actualtaupa").set_text("%.4g" % self.afterfitparams[0])
		self.builder.get_object("actualtauna").set_text("%.4g" % self.afterfitparams[1])
		self.builder.get_object("actualn1a").set_text("%.4g" % self.afterfitparams[2])
		self.builder.get_object("actualp1a").set_text("%.4g" % self.afterfitparams[3])
		self.builder.get_object("actualNAa").set_text("%.4g" % self.afterfitparams[4])
		self.builder.get_object("actualtaupb").set_text("%.4g" % self.beforefitparams[0])
		self.builder.get_object("actualtaunb").set_text("%.4g" % self.beforefitparams[1])
		self.builder.get_object("actualn1b").set_text("%.4g" % self.beforefitparams[2])
		self.builder.get_object("actualp1b").set_text("%.4g" % self.beforefitparams[3])
		self.builder.get_object("actualNAb").set_text("%.4g" % self.beforefitparams[4])
		self.builder.get_object("bfitrmin").set_text("%.4g" % self.bfitrmin)
		self.builder.get_object("bfitrmax").set_text("%.4g" % self.bfitrmax)
		self.builder.get_object("afitrmin").set_text("%.4g" % self.afitrmin)
		self.builder.get_object("afitrmax").set_text("%.4g" % self.afitrmax)
		self.builder.get_object("editfitwindow").show()

	def fitdefaultsbtnclicked(self, widget):
		self.startguessa=self.defaultstartguess
		self.startguessb=self.defaultstartguess
		self.builder.get_object("taupb").set_text("%.4g" % self.startguessb[0])
		self.builder.get_object("taunb").set_text("%.4g" % self.startguessb[1])
		self.builder.get_object("n1b").set_text("%.4g" % self.startguessb[2])
		self.builder.get_object("p1b").set_text("%.4g" % self.startguessb[3])
		self.builder.get_object("NAb").set_text("%.4g" % self.startguessb[4])
		self.builder.get_object("taupa").set_text("%.4g" % self.startguessa[0])
		self.builder.get_object("tauna").set_text("%.4g" % self.startguessa[1])
		self.builder.get_object("n1a").set_text("%.4g" % self.startguessa[2])
		self.builder.get_object("p1a").set_text("%.4g" % self.startguessa[3])
		self.builder.get_object("NAa").set_text("%.4g" % self.startguessa[4])
		self.beforeplotcounter=0
		self.afterplotcounter=0

	def editfitcancelbtnclicked(self, widget):
		self.builder.get_object("editfitwindow").hide()

	def editfitokbtnclicked(self, widget):
		self.bfitrmin=float(self.builder.get_object("bfitrmin").get_text())
		self.bfitrmax=float(self.builder.get_object("bfitrmax").get_text())
		self.afitrmin=float(self.builder.get_object("afitrmin").get_text())
		self.afitrmax=float(self.builder.get_object("afitrmax").get_text())
		self.startguessb=np.array([float(self.builder.get_object("taupb").get_text()),float(self.builder.get_object("taunb").get_text()),float(self.builder.get_object("n1b").get_text()),float(self.builder.get_object("p1b").get_text()),float(self.builder.get_object("NAb").get_text())],np.float64)
		self.startguessa=np.array([float(self.builder.get_object("taupa").get_text()),float(self.builder.get_object("tauna").get_text()),float(self.builder.get_object("n1a").get_text()),float(self.builder.get_object("p1a").get_text()),float(self.builder.get_object("NAa").get_text())],np.float64)

		self.plotgraph1(widget)
		self.builder.get_object("editfitwindow").hide()

	def compareplmapsbtnclicked(self, widget):
		#PLOT MAPS
		self.plotcompare(widget)
		self.builder.get_object("comparewindow").show()
		
	def plotcompare(self, widget):
		
		self.constants=[float(self.builder.get_object("vthermaltxt").get_text()), float(self.builder.get_object("sigmanitxt").get_text()), float(self.builder.get_object("p1itxt").get_text()), float(self.builder.get_object("sigmapitxt").get_text()), float(self.builder.get_object("sigmanbtxt").get_text()), float(self.builder.get_object("n1btxt").get_text()), float(self.builder.get_object("sigmapbtxt").get_text())]
		l_f = LogFormatter(10, labelOnlyBase=False)
		if self.compcounter==0:
			self.compfigure1 = Figure(figsize=(6,4), dpi=72)  
			self.compfigure2=Figure(figsize=(6,4), dpi=72)
			self.compfigure3=Figure(figsize=(6,4), dpi=72)
			self.compfigure4=Figure(figsize=(6,4), dpi=72)
			self.compfigure5=Figure(figsize=(6,4), dpi=72)			
			self.compfigure6=Figure(figsize=(6,4), dpi=72)			
			
		if self.compcounter>0:
			self.compaxis1.clear()
			self.compaxis2.clear()
			self.compaxis3.clear()
			self.compaxis4.clear()
			self.compaxis5.clear()
			self.compaxis6.clear()			
			self.compfigure1.clear()
			self.compfigure2.clear()
			self.compfigure3.clear()
			self.compfigure4.clear()
			self.compfigure5.clear()
			self.compfigure6.clear()			
			self.builder.get_object("individualcompare").remove(self.compcanvas1)
			self.builder.get_object("differencescompare").remove(self.compcanvas2)
			self.builder.get_object("genlevelcompare").remove(self.compcanvas3)
			self.builder.get_object("meanrangecompare").remove(self.compcanvas4)
			self.builder.get_object("indgencompare").remove(self.compcanvas5)
			self.builder.get_object("differencescompareindgen").remove(self.compcanvas6)
			
			

		self.compaxis1=self.compfigure1.add_subplot(111, aspect='equal')
		self.compaxis2=self.compfigure2.add_subplot(111, aspect='equal')
		self.compaxis3=self.compfigure3.add_subplot(111, aspect='equal')
		self.compaxis4=self.compfigure4.add_subplot(111, aspect='equal')
		self.compaxis5=self.compfigure5.add_subplot(111, aspect='equal')
		self.compaxis6=self.compfigure6.add_subplot(111, aspect='equal')
		
				
		if self.ironmaptype=="linear":
			imageironindividual=self.compaxis1.imshow(self.ironconcmatrixindividual, cmap=self.cmap, vmin=self.ironmin, vmax=self.ironmax)
			cbarindividual=self.compfigure1.colorbar(imageironindividual, fraction=0.045, extend='both')
			imageironindgen=self.compaxis5.imshow(self.ironconcmatrixindgen, cmap=self.cmap, vmin=self.ironmin, vmax=self.ironmax)
			cbarindgen=self.compfigure5.colorbar(imageironindgen, fraction=0.045, extend='both')			
			imageirongenlevel=self.compaxis3.imshow(self.ironconcmatrixgenlevel, cmap=self.cmap, vmin=self.ironmin, vmax=self.ironmax)
			cbargenlevel=self.compfigure3.colorbar(imageirongenlevel, fraction=0.045, extend='both')
			imageironmeanC=self.compaxis4.imshow(self.ironconcmatrixmeanC, cmap=self.cmap, vmin=self.ironmin, vmax=self.ironmax)
			cbarmeanC=self.compfigure4.colorbar(imageironmeanC, fraction=0.045, extend='both')			
			
		elif self.ironmaptype=="logarithmic":
			imageironindividual=self.compaxis1.imshow(self.ironconcmatrixindividual, cmap=self.cmap, norm=LogNorm(vmin=self.ironmin, vmax=self.ironmax))
			cbarindividual=self.compfigure1.colorbar(imageironindividual, format=l_f,fraction=0.045, extend='both')
			imageironindgen=self.compaxis5.imshow(self.ironconcmatrixindgen, cmap=self.cmap, norm=LogNorm(vmin=self.ironmin, vmax=self.ironmax))
			cbarindgen=self.compfigure5.colorbar(imageironindgen, format=l_f,fraction=0.045, extend='both')			
			imageirongenlevel=self.compaxis3.imshow(self.ironconcmatrixgenlevel, cmap=self.cmap, norm=LogNorm(vmin=self.ironmin, vmax=self.ironmax))
			cbargenlevel=self.compfigure3.colorbar(imageirongenlevel, format=l_f,fraction=0.045, extend='both')
			imageironmeanC=self.compaxis4.imshow(self.ironconcmatrixmeanC, cmap=self.cmap, norm=LogNorm(vmin=self.ironmin, vmax=self.ironmax))
			cbarmeanC=self.compfigure4.colorbar(imageironmeanC, format=l_f,fraction=0.045, extend='both')


		cbarindividual.set_label(r'[Fe$_{i}$]  (cm$^{-3})$')
		cbarindgen.set_label(r'[Fe$_{i}$]  (cm$^{-3})$')		
		cbargenlevel.set_label(r'[Fe$_{i}$]  (cm$^{-3})$')
		cbarmeanC.set_label(r'[Fe$_{i}$]  (cm$^{-3})$')

		imagedifferences=self.compaxis2.imshow(self.diffmatrix, cmap=self.cmap, vmin=0, vmax=self.ddnlimit)
		cbardifferences=self.compfigure2.colorbar(imagedifferences, fraction=0.045, extend='both')
		cbardifferences.set_label(r'Difference in injection levels')
		imagedifferencesindgen=self.compaxis6.imshow(self.diffmatrixindgen, cmap=self.cmap, vmin=0, vmax=self.ddnlimit)
		cbardifferencesindgen=self.compfigure6.colorbar(imagedifferencesindgen, fraction=0.045, extend='both')
		cbardifferencesindgen.set_label(r'Difference in injection levels')		
		self.compcanvas1=FigureCanvasGTKAgg(self.compfigure1)
		self.compcanvas1.show()
		self.compcanvas2=FigureCanvasGTKAgg(self.compfigure2)
		self.compcanvas2.show()
		self.compcanvas3=FigureCanvasGTKAgg(self.compfigure3)
		self.compcanvas3.show()
		self.compcanvas4=FigureCanvasGTKAgg(self.compfigure4)
		self.compcanvas4.show()
		self.compcanvas5=FigureCanvasGTKAgg(self.compfigure5)
		self.compcanvas5.show()
		self.compcanvas6=FigureCanvasGTKAgg(self.compfigure6)
		self.compcanvas6.show()			
		
		self.compaxis1.set_xlim(left=0, right=1000)
		self.compaxis1.set_ylim(top=0, bottom=1000)
		self.compaxis2.set_xlim(left=0, right=1000)
		self.compaxis2.set_ylim(top=0, bottom=1000)
		self.compaxis3.set_xlim(left=0, right=1000)
		self.compaxis3.set_ylim(top=0, bottom=1000)
		self.compaxis4.set_xlim(left=0, right=1000)
		self.compaxis4.set_ylim(top=0, bottom=1000)
		self.compaxis5.set_xlim(left=0, right=1000)
		self.compaxis5.set_ylim(top=0, bottom=1000)
		self.compaxis6.set_xlim(left=0, right=1000)
		self.compaxis6.set_ylim(top=0, bottom=1000)		
		
		self.builder.get_object("individualcompare").pack_start(self.compcanvas1, True, True)
		self.builder.get_object("differencescompare").pack_start(self.compcanvas2, True, True)
		self.builder.get_object("genlevelcompare").pack_start(self.compcanvas3, True, True)
		self.builder.get_object("meanrangecompare").pack_start(self.compcanvas4, True, True)
		self.builder.get_object("indgencompare").pack_start(self.compcanvas5, True, True)
		self.builder.get_object("differencescompareindgen").pack_start(self.compcanvas6, True, True)
		self.compcanvas1.mpl_connect('motion_notify_event', self.graphscrollcomp)
		self.compcanvas2.mpl_connect('motion_notify_event', self.graphscrollcomp)
		self.compcanvas3.mpl_connect('motion_notify_event', self.graphscrollcomp)
		self.compcanvas4.mpl_connect('motion_notify_event', self.graphscrollcomp)		
		self.compcanvas5.mpl_connect('motion_notify_event', self.graphscrollcomp)
		self.compcanvas6.mpl_connect('motion_notify_event', self.graphscrollcomp)		
		#self.builder.get_object("pltoolbar").set_sensitive(True)
		#self.builder.get_object("editcolorbarbtn").set_sensitive(True)
		self.compcounter+=1

	def closecomparebtnclicked(self, widget):
		self.builder.get_object("comparewindow").hide()

	def zoomincomparebtnclicked(self, widget):
		xlims=self.compaxis1.get_xlim()
		ylims=self.compaxis1.get_ylim()
		xchange=abs(xlims[1]-xlims[0])*0.1
		ychange=abs(ylims[1]-ylims[0])*0.1
		self.compaxis1.set_xlim(left=xlims[0]+xchange, right=xlims[1]-xchange)
		self.compaxis1.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]-ychange)
		self.compaxis2.set_xlim(left=xlims[0]+xchange, right=xlims[1]-xchange)
		self.compaxis2.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]-ychange)
		self.compaxis3.set_xlim(left=xlims[0]+xchange, right=xlims[1]-xchange)
		self.compaxis3.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]-ychange)
		self.compaxis4.set_xlim(left=xlims[0]+xchange, right=xlims[1]-xchange)
		self.compaxis4.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]-ychange)
		self.builder.get_object("individualcompare").remove(self.compcanvas1)
		self.builder.get_object("differencescompare").remove(self.compcanvas2)
		self.builder.get_object("genlevelcompare").remove(self.compcanvas3)
		self.builder.get_object("meanrangecompare").remove(self.compcanvas4)
		self.builder.get_object("individualcompare").pack_start(self.compcanvas1, True, True)
		self.builder.get_object("differencescompare").pack_start(self.compcanvas2, True, True)
		self.builder.get_object("genlevelcompare").pack_start(self.compcanvas3, True, True)
		self.builder.get_object("meanrangecompare").pack_start(self.compcanvas4, True, True)		
		
	def zoomoutcomparebtnclicked(self, widget):
		xlims=self.compaxis1.get_xlim()
		ylims=self.compaxis1.get_ylim()
		xchange=abs(xlims[1]-xlims[0])*0.1
		xchange=xchange/0.8
		ychange=abs(ylims[1]-ylims[0])*0.1
		ychange=ychange/0.8
		self.compaxis1.set_xlim(left=xlims[0]-xchange, right=xlims[1]+xchange)
		self.compaxis1.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]+ychange)
		self.compaxis2.set_xlim(left=xlims[0]-xchange, right=xlims[1]+xchange)
		self.compaxis2.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]+ychange)
		self.compaxis3.set_xlim(left=xlims[0]-xchange, right=xlims[1]+xchange)
		self.compaxis3.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]+ychange)
		self.compaxis4.set_xlim(left=xlims[0]-xchange, right=xlims[1]+xchange)
		self.compaxis4.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]+ychange)
		self.builder.get_object("individualcompare").remove(self.compcanvas1)
		self.builder.get_object("differencescompare").remove(self.compcanvas2)
		self.builder.get_object("genlevelcompare").remove(self.compcanvas3)
		self.builder.get_object("meanrangecompare").remove(self.compcanvas4)
		self.builder.get_object("individualcompare").pack_start(self.compcanvas1, True, True)
		self.builder.get_object("differencescompare").pack_start(self.compcanvas2, True, True)
		self.builder.get_object("genlevelcompare").pack_start(self.compcanvas3, True, True)
		self.builder.get_object("meanrangecompare").pack_start(self.compcanvas4, True, True)	

       	def panleftcomparebtnclicked(self, widget):
		xlims=self.compaxis1.get_xlim()
		xchange=abs(xlims[1]-xlims[0])*0.1
		self.compaxis1.set_xlim(left=xlims[0]-xchange, right=xlims[1]-xchange)
		self.compaxis2.set_xlim(left=xlims[0]-xchange, right=xlims[1]-xchange)
		self.compaxis3.set_xlim(left=xlims[0]-xchange, right=xlims[1]-xchange)
		self.compaxis4.set_xlim(left=xlims[0]-xchange, right=xlims[1]-xchange)		
		self.builder.get_object("individualcompare").remove(self.compcanvas1)
		self.builder.get_object("differencescompare").remove(self.compcanvas2)
		self.builder.get_object("genlevelcompare").remove(self.compcanvas3)
		self.builder.get_object("meanrangecompare").remove(self.compcanvas4)
		self.builder.get_object("individualcompare").pack_start(self.compcanvas1, True, True)
		self.builder.get_object("differencescompare").pack_start(self.compcanvas2, True, True)
		self.builder.get_object("genlevelcompare").pack_start(self.compcanvas3, True, True)
		self.builder.get_object("meanrangecompare").pack_start(self.compcanvas4, True, True)

       	def panrightcomparebtnclicked(self, widget):
		xlims=self.compaxis1.get_xlim()
		xchange=abs(xlims[1]-xlims[0])*0.1
		self.compaxis1.set_xlim(left=xlims[0]+xchange, right=xlims[1]+xchange)
		self.compaxis2.set_xlim(left=xlims[0]+xchange, right=xlims[1]+xchange)
		self.compaxis3.set_xlim(left=xlims[0]+xchange, right=xlims[1]+xchange)
		self.compaxis4.set_xlim(left=xlims[0]+xchange, right=xlims[1]+xchange)
		self.builder.get_object("individualcompare").remove(self.compcanvas1)
		self.builder.get_object("differencescompare").remove(self.compcanvas2)
		self.builder.get_object("genlevelcompare").remove(self.compcanvas3)
		self.builder.get_object("meanrangecompare").remove(self.compcanvas4)
		self.builder.get_object("individualcompare").pack_start(self.compcanvas1, True, True)
		self.builder.get_object("differencescompare").pack_start(self.compcanvas2, True, True)
		self.builder.get_object("genlevelcompare").pack_start(self.compcanvas3, True, True)
		self.builder.get_object("meanrangecompare").pack_start(self.compcanvas4, True, True)

	def panupcomparebtnclicked(self, widget):
		ylims=self.compaxis1.get_ylim()
		ychange=abs(ylims[1]-ylims[0])*0.1
		self.compaxis1.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]-ychange)
		self.compaxis2.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]-ychange)
		self.compaxis3.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]-ychange)
		self.compaxis4.set_ylim(top=ylims[1]-ychange, bottom=ylims[0]-ychange)
		self.builder.get_object("individualcompare").remove(self.compcanvas1)
		self.builder.get_object("differencescompare").remove(self.compcanvas2)
		self.builder.get_object("genlevelcompare").remove(self.compcanvas3)
		self.builder.get_object("meanrangecompare").remove(self.compcanvas4)
		self.builder.get_object("individualcompare").pack_start(self.compcanvas1, True, True)
		self.builder.get_object("differencescompare").pack_start(self.compcanvas2, True, True)
		self.builder.get_object("genlevelcompare").pack_start(self.compcanvas3, True, True)
		self.builder.get_object("meanrangecompare").pack_start(self.compcanvas4, True, True)		

	def pandowncomparebtnclicked(self, widget):
		ylims=self.compaxis1.get_ylim()
		ychange=abs(ylims[1]-ylims[0])*0.1
		self.compaxis1.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]+ychange)
		self.compaxis2.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]+ychange)
		self.compaxis3.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]+ychange)
		self.compaxis4.set_ylim(top=ylims[1]+ychange, bottom=ylims[0]+ychange)		
		self.builder.get_object("individualcompare").remove(self.compcanvas1)
		self.builder.get_object("differencescompare").remove(self.compcanvas2)
		self.builder.get_object("genlevelcompare").remove(self.compcanvas3)
		self.builder.get_object("meanrangecompare").remove(self.compcanvas4)
		self.builder.get_object("individualcompare").pack_start(self.compcanvas1, True, True)
		self.builder.get_object("differencescompare").pack_start(self.compcanvas2, True, True)
		self.builder.get_object("genlevelcompare").pack_start(self.compcanvas3, True, True)
		self.builder.get_object("meanrangecompare").pack_start(self.compcanvas4, True, True)

	def refreshcomparebtnclicked(self, widget):
		self.plotcompare(widget)

	def graphscrollcomp(self, event):
		if event.x!=None and event.y!=None and event.xdata!=None and event.ydata!=None:

			if self.curid2!=None:
				self.builder.get_object("comparestatusbar").remove_message(self.plconid2, self.curid2)
				
			self.curid2=self.builder.get_object("comparestatusbar").push(self.plconid2, 'x=%d, y=%d, taub=%.3g, taua=%.3g, ironindividualCfits=%.3g, deltandifferencefits=%.3g, ironindividualCtaugen=%.3g, deltandifferencetaugen=%.3g, irongenlevelC=%.3g, ironlastmeanC=%.3g'%(int(round(event.xdata)), int(round(event.ydata)), self.taubefore[int(round(event.ydata))][int(round(event.xdata))], self.tauafter[int(round(event.ydata))][int(round(event.xdata))], self.ironconcmatrixindividual[int(round(event.ydata))][int(round(event.xdata))], self.diffmatrix[int(round(event.ydata))][int(round(event.xdata))], self.ironconcmatrixindgen[int(round(event.ydata))][int(round(event.xdata))], self.diffmatrixindgen[int(round(event.ydata))][int(round(event.xdata))], self.ironconcmatrixgenlevel[int(round(event.ydata))][int(round(event.xdata))], self.ironconcmatrixmeanC[int(round(event.ydata))][int(round(event.xdata))]))		

if __name__ == "__main__":
	app = MyApp()
	mainloop=gtk.main()
